from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from inventory_app.models import Finished_Goods_Request, Finished_Goods_Stock, Damaged_Goods_Stock
from inventory_app.forms import Finished_Goods_RequestForm, Finished_Goods_StockForm, Damaged_Goods_StockForm
from production_app.forms import water_category_Form, finishedwaterForm, DamagedForm, damaged_goods_Form, \
    update_damaged_goods_Form
from production_app.models import water_Finished_goods_category, water_Finished_Goods, damaged_Goods, \
    Damaged_good_category
from register_app.models import CustomUser, MenuPermissions
from django.core.paginator import Paginator
from django.db.models import F, Q
from django.db.models import Sum
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.template.loader import get_template
from django.db.models import DateField, Value
from django.db.models.functions import Cast
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from xhtml2pdf import pisa
from django.urls import reverse


# Create your views here.

def view_finished_water_good(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    categories =water_Finished_goods_category.objects.all()
    view_ca=water_Finished_Goods.objects.all()


    # Pagination logic
    paginator = Paginator(view_ca, 5)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request,'production/view_finished_goods.html',{'view':page_obj,'categories':categories,'allocated_menus':allocated_menus})


def addwater_category(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    form = water_category_Form()
    if request.method == 'POST':
        form = water_category_Form(request.POST)
        print("Request")
        if form.is_valid():
            print("Success")
            form.save()
            messages.success(request, 'Successfull')

            return redirect('view_finished_water_good')

    return render(request,'production/create_water_category.html',{'form':form,'allocated_menus':allocated_menus})



def create_water(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    if request.method == 'POST':
        form = finishedwaterForm(request.POST, request.FILES)  # Ensure request.FILES is used for image upload
        if form.is_valid():
            form.save()
            return redirect('view_finished_water_good')  # Replace with the correct URL or view name
    else:
        form = finishedwaterForm()


    return render(request, 'production/create_water.html', {'form': form, 'allocated_menus':allocated_menus})


def update_finished_goods(request, id):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    mem = get_object_or_404(water_Finished_Goods, id=id)

    if request.method == 'POST':
        form = finishedwaterForm(request.POST,request.FILES,instance=mem)
        if form.is_valid():
            form.save()
            return redirect('view_finished_water_good')  # Redirect to the list view or any relevant view
    else:
        form = finishedwaterForm(instance=mem)


    return render(request, 'production/update_water.html', {'form': form,'allocated_menus':allocated_menus})



def delete_goods(request, id):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Fetch the object or return a 404 error if not found
    dtl = get_object_or_404(water_Finished_Goods, id=id)

    if request.method == "POST":
        dtl.delete()
        return redirect('view_finished_water_good')

    # Render the confirmation page for GET requests

    return render(request, 'production/delete_water.html', {'dtl': dtl,'allocated_menus':allocated_menus})




def add_damaged_good_category(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)


    if request.method == "POST":
        form = DamagedForm(request.POST)
        if form.is_valid():
            form.save()

            messages.success(request, 'Successfull')
            return redirect('damaged_goods')
    else:
        form = DamagedForm(request.POST)
    return render(request, 'production/damaged_good_category.html', {'form': form, 'allocated_menus': allocated_menus})

def damaged_goods(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    cc = Damaged_good_category.objects.all()
    dd=damaged_Goods.objects.all()

    # Pagination logic
    paginator = Paginator(dd, 5)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'production/damaged_goods.html', {'allocated_menus':allocated_menus,'dd':page_obj, 'cc': cc})


def add_damagedgoods(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    if request.method == 'POST':
        form = damaged_goods_Form(request.POST,request.FILES)
        if form.is_valid():
            form.save()
            return redirect('damaged_goods')  # Replace 'Route' with the actual view name or URL pattern
    else:
        form = damaged_goods_Form()


    return render(request, 'production/add_damaged_goods.html', {'form': form,'allocated_menus': allocated_menus})



def damage_delete(request,id):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    details = get_object_or_404(damaged_Goods,id=id)

    if request.method == 'POST':
        try:
            details.delete()
            return redirect('damaged_goods')
        except Exception as e:
            print(e)

    return render(request,'production/damage_delete.html',{'allocated_menus': allocated_menus})


def update_damage(request,id):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)


    details = get_object_or_404(damaged_Goods, id=id)

    if request.method == 'POST':
        form = update_damaged_goods_Form(request.POST,instance=details)
        if form.is_valid():
            form.save()
            return redirect('damaged_goods')

    else:
        form = update_damaged_goods_Form(instance=details)

    return render(request,'production/update_damage.html',{'allocated_menus':allocated_menus,'form':form})



def request_messages(request):
    # Fetch the current logged-in user
    use = request.user

    # Get the user's permissions for menus
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)



    # Fetch only the requests with status 'Pending'
    data = Finished_Goods_Request.objects.filter(status='Pending').order_by('-date')

    # Check if there are any pending requests

    # Pagination logic
    paginator = Paginator(data, 5)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Render the page with the data
    return render(request, 'production/production_request_messages.html', {
        'data': page_obj,
        'allocated_menus': allocated_menus
    })



def request_messages_detail(request, id):
    # Fetch menus and counts
    use = request.user
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    request_data = get_object_or_404(Finished_Goods_Request, pk=id)
    page = request.GET.get('page', 1)  # Capture current page number

    if request.method == 'POST':
        action_type = request.POST.get('action_type')
        response_text = request.POST.get('response')

        if action_type == 'accept':
            request_data.status = 'Accepted'
            request_data.response = response_text
            request_data.save()
            return redirect('request_messages')

        elif action_type == 'decline':
            if response_text:
                request_data.status = 'Declined'
                request_data.response = response_text
                request_data.save()
                return redirect('request_messages')
            else:
                error_message = "Please provide a reason for declining"
                return render(request, 'production/production_request_messages.html', {
                    'data': request_data,
                    'allocated_menus': allocated_menus,
                    'error_message': error_message,
                    'page': page,  # Pass page number to the template
                })

    return render(request, 'production/production_request_messages_in_detail.html', {
        'data': request_data,
        'allocated_menus': allocated_menus,
        'page': page,  # Pass page number to the template

    })


def search(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Fetch all categories
    categories = Damaged_good_category.objects.all()

    # Initialize the queryset for damaged goods
    damaged_goods_list = damaged_Goods.objects.all()

    # Handle search requests
    if request.method == 'GET':
        damaged_name = request.GET.get('name', None)
        damaged_category = request.GET.get('category', None)

        # Build filters
        filters = Q()

        # If the search name is provided, filter by the name in damaged_Goods
        if damaged_name:
            filters &= Q(name__icontains=damaged_name)

        # If a valid category is selected, filter by the category
        if damaged_category:
            filters &= Q(category_id=damaged_category)  # Use category_id to filter

        # Apply the filters to the queryset if filters are provided
        if filters:
            damaged_goods_list = damaged_Goods.objects.filter(filters)

    # Prepare context with the filtered or full list of damaged goods
    context = {
        'dd': damaged_goods_list,  # This is the queryset for the table in the HTML
        'categories': categories,
        'allocated_menus':allocated_menus
    }

    return render(request, 'production/damaged_goods.html', context)


def finishedgoods_production_search(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Get distinct categories and finished goods
    categories = water_Finished_goods_category.objects.all()
    finished_goods = water_Finished_Goods.objects.prefetch_related('stocks').all()

    # Apply search filters if form is submitted
    if request.method == 'POST':
        category = request.POST.get('category')
        name = request.POST.get('name')
        filters = Q()

        # Apply category filter if selected
        if category and category.isdigit():
            filters &= Q(category_id=int(category))

        # Apply name filter if provided
        if name:
            filters &= Q(name__icontains=name)

        # Apply the filters to finished_goods queryset
        if filters:
            finished_goods = water_Finished_Goods.objects.filter(filters)
        else:
            finished_goods = water_Finished_Goods.objects.none()

    # Paginate the finished goods results
    paginator = Paginator(finished_goods, 5)  # Show 5 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'view': page_obj,  # Paginated finished goods
        'categories': categories,
        'allocated_menus': allocated_menus
    }

    return render(request, 'production/view_finished_goods.html', context)
def update_finished_goods_stocks(request, id):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)


    finished_goods = get_object_or_404(water_Finished_Goods, id=id)

    if request.method == 'POST':
        form = Finished_Goods_StockForm(request.POST)
        if form.is_valid():
            stock_entry = form.save(commit=False)
            stock_entry.finished_goods = finished_goods  # Link stock entry to finished_goods
            stock_entry.save()  # Save the new stock entry

            # Add the new stock to the existing total stock
            new_stock = stock_entry.stock
            finished_goods.total_stock = F('total_stock') + new_stock
            finished_goods.save(update_fields=["total_stock"])  # Save the updated total stock

            return redirect('view_finished_water_good')  # Redirect to the list view
    else:
        form = Finished_Goods_StockForm()

    return render(request, 'production/add_finishedgoods_stock.html', {
        'form': form,
        'finished_goods': finished_goods,
        'allocated_menus': allocated_menus
    })


def finishedgoods_production_stock_history(request, id):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Get initial queryset for finished goods based on the provided ID
    finished_good = Finished_Goods_Stock.objects.filter(finished_goods_id=id).order_by('-date','-time')

    # Get filter type and query date from the request
    filter_type = request.GET.get('filter')
    query_date = request.GET.get('query')  # Expected format: YYYY-MM-DD
    query_date_end = request.GET.get('query_end')  # Expected format: YYYY-MM-DD

    is_filtered = False

    # Apply filtering if filter type is 'date' and query_date is provided
    if filter_type == 'date' and query_date and query_date_end:
        finished_good = finished_good.filter(date__range=[query_date, query_date_end])
        is_filtered = True  # Flag to indicate that a filter has been applied

    # Pagination logic
    paginator = Paginator(finished_good, 5)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Calculate total stock for filtered results
    total_stock = sum(item.stock for item in finished_good)

    # Prepare context for rendering template
    context = {
        'data': page_obj,
        'allocated_menus': allocated_menus,
        'id': id,
        'filter_type': filter_type,
        'query_date': query_date,
        'is_filtered': is_filtered,
        'total_stock': total_stock,
        'query_date_end': query_date_end
    }

    return render(request, 'production/view_finished_goods_history.html', context)



def finished_goods_production_history_pdf(request, id):
    # Retrieve and validate the query date parameters
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date_end else None

    # Apply date filter only if both parsed dates are valid
    if parsed_date and parsed_date_end:
        data = Finished_Goods_Stock.objects.filter(
            finished_goods_id=id,
            date__isnull=False,
            date__range=(parsed_date, parsed_date_end) # Filter within the date range
        ).order_by('-date','-time')
    else:
        data = Finished_Goods_Stock.objects.filter(finished_goods_id=id).order_by('-date','-time')

    # Retrieve the first entry for dynamic filename generation
    view = data.first()
    filename = f"{view.finished_goods.category.category_name}_production_report.pdf" if view else "production_report.pdf"

    # Prepare context for rendering the PDF template
    context = {'view': data, 'name': view}

    # Load and render the template to HTML
    template = get_template('production/finishedgoods_production_stock_pdf.html')
    html = template.render(context)

    # Create the HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Generate the PDF
    pdf_status = pisa.CreatePDF(html, dest=response)

    if pdf_status.err:
        return HttpResponse("Error generating PDF", status=400)

    return response


def finished_goods_production_history_excel(request, id):
    # Retrieve and validate the query date parameters
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date_end else None

    # Apply date filter only if both parsed dates are valid
    if parsed_date and parsed_date_end:
        finished_good = Finished_Goods_Stock.objects.filter(
            finished_goods_id=id,
            date__isnull=False
        ).order_by('-date','-time') .filter(
            date__range=([parsed_date, parsed_date_end]) # Filter by date range
        )
    else:
        finished_good = Finished_Goods_Stock.objects.filter(finished_goods_id=id).order_by('-date','-time')

    # Get the category name for the filename
    data = finished_good.first()  # Use first entry to get category info
    filename = f"{data.finished_goods.category.category_name}_production_report.xlsx" if data else "production_report.xlsx"

    # Create an Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Production Report"

    # Write the title row at the top
    title = f"{data.finished_goods.category.category_name} Production Report" if data else "Production Report"
    ws.append([title])  # Title row
    ws.append([])  # Blank row for line space
    ws.append([])  # Second blank row for line space

    # Write the header row (these are your "bold" headers in Excel)
    headers = ['Category', 'Name', 'Size', 'Stock', 'Remarks','Date']
    ws.append(headers)

    # Set header row font to bold
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)  # Header is in row 4 after title rows
        cell.font = header_font

    # Iterate through the filtered data and write each row to the Excel file
    for i in finished_good:
        ws.append([
            i.finished_goods.category.category_name,
            i.finished_goods.name,
            i.finished_goods.size,
            i.stock,
            i.remarks,
            i.date.strftime('%Y-%m-%d') if i.date else 'N/A'  # Formatting date as needed
        ])

    # Prepare the response as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save the workbook to the response
    wb.save(response)

    return response


def damagedgoods_production_stock_history(request, id):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Get initial queryset for damaged goods based on the provided ID
    damaged_goods = Damaged_Goods_Stock.objects.filter(damaged_id=id).order_by('-date','-time')


    # Get filter type and query date from the request
    filter_type = request.GET.get('filter')
    query_date = request.GET.get('query')  # Expected format: YYYY-MM-DD
    query_date_end = request.GET.get('query_end')
    is_filtered = False

    # Apply filtering if filter type is 'date' and query_date is provided
    if filter_type == 'date' and query_date and query_date_end:
        damaged_goods = damaged_goods.filter(date__range=[query_date, query_date_end])
        is_filtered = True  # Flag to indicate that a filter has been applied

    # Pagination logic
    paginator = Paginator(damaged_goods, 5)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Calculate total stock for filtered results
    total_stock = sum(item.stock for item in damaged_goods)

    # Prepare context for rendering template
    context = {
        'id': id,
        'filter_type': filter_type,
        'query_date': query_date,
        'query_date_end': query_date_end,
        'is_filtered': is_filtered,  # Pass flag to template
        'allocated_menus': allocated_menus,
        'data': page_obj,
        'total_stock': total_stock
    }

    return render(request, 'production/damaged_production_stock_history.html', context)


def update_damaged_goods_stocks(request, id):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    damaged_goods = get_object_or_404(damaged_Goods, id=id)

    if request.method == 'POST':
        form = Damaged_Goods_StockForm(request.POST)
        if form.is_valid():
            stock_entry = form.save(commit=False)
            stock_entry.damaged = damaged_goods  # Set the raw material
            stock_entry.save()  # Save the new stock entry

            # Update the total stock for the raw material
            total_stock = damaged_goods.stocks.aggregate(total=Sum('stock'))['total'] or 0
            damaged_goods.total_stock = total_stock
            damaged_goods.save()  # Save the updated total stock

            return redirect('damaged_goods')  # Redirect to the list view
    else:
        form = Damaged_Goods_StockForm()

    return render(request, 'production/add_damage_stock.html', {
        'form': form,
        'damaged_goods': damaged_goods,
        'allocated_menus': allocated_menus
    })

def damaged_goods_production_history_pdf(request, id):
    # Retrieve and validate the query date parameters
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date_end else None

    # Apply date filter only if both parsed dates are valid
    if parsed_date and parsed_date_end:
        data = Damaged_Goods_Stock.objects.filter(
            damaged_id=id,
            date__isnull=False,
            date__range=(parsed_date, parsed_date_end)  # Filter within the date range
        ).order_by('-date','-time')
    else:
        data = Damaged_Goods_Stock.objects.filter(damaged_id=id).order_by('-date','-time')

    # Retrieve the first entry for dynamic filename generation
    view = data.first()
    filename = f"{view.damaged.category.category_name}_production_report.pdf" if view else "production_report.pdf"

    # Prepare context for rendering the PDF template
    context = {'view': data, 'name': view}

    # Load and render the template to HTML
    template = get_template('production/damagedgoods_production_history_pdf.html')
    html = template.render(context)

    # Create the HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Generate the PDF
    pdf_status = pisa.CreatePDF(html, dest=response)

    if pdf_status.err:
        return HttpResponse("Error generating PDF", status=400)

    return response

def damaged_goods_production_history_excel(request, id):
    # Retrieve and validate the query date parameters
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date_end else None

    # Apply date filter only if both parsed dates are valid
    if parsed_date and parsed_date_end:
        damaged_good = Damaged_Goods_Stock.objects.filter(
            damaged_id=id,
            date__isnull=False
        ).order_by('-date','-time') .filter(
            date__range=([parsed_date, parsed_date_end])  # Filter by date range
        )
    else:
        damaged_good = Damaged_Goods_Stock.objects.filter(damaged_id=id).order_by('-date','-time')

    # Get the category name for the filename
    data = damaged_good.first()  # Use first entry to get category info
    filename = f"{data.damaged.category.category_name}_production_report.xlsx" if data else "production_report.xlsx"

    # Create an Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Production Report"

    # Write the title row at the top
    title = f"{data.damaged.category.category_name} Production Report" if data else "Production Report"
    ws.append([title])  # Title row
    ws.append([])  # Blank row for line space
    ws.append([])  # Second blank row for line space

    # Write the header row (these are your "bold" headers in Excel)
    headers = ['Category', 'Name', 'Description', 'Stock', 'Date']
    ws.append(headers)

    # Set header row font to bold
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)  # Header is in row 4 after title rows
        cell.font = header_font

    # Iterate through the filtered data and write each row to the Excel file
    for i in damaged_good:
        ws.append([
            i.damaged.category.category_name,
            i.damaged.name,
            i.damaged.description,
            i.stock,
            i.date.strftime('%Y-%m-%d') if i.date else 'N/A'  # Formatting date as needed
        ])

    # Prepare the response as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save the workbook to the response
    wb.save(response)

    return response


def finishedgoods_message_request_history(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    data = Finished_Goods_Request.objects.all().order_by('-date')

    # Get filter type and query date from the request
    filter_type = request.GET.get('filter')
    query_date = request.GET.get('query')  # Expected format: YYYY-MM-DD
    query_date_end = request.GET.get('query_end')
    is_filtered = False

    # Apply filtering if filter type is 'date' and query_date is provided
    if filter_type == 'date' and query_date and query_date_end:
        data = data.annotate(date_only=Cast('date', DateField())).filter(date_only__range=[query_date, query_date_end])
        is_filtered = True  # Flag to indicate that a filter has been applied

    elif query_date:
        data = data.annotate(date_only=Cast('date', DateField())).filter(date_only=query_date)
        is_filtered = True  # Flag to indicate that a filter has been applied

    # Pagination logic
    paginator = Paginator(data, 5)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'data': page_obj,
        'allocated_menus': allocated_menus,
        'is_filtered': is_filtered,
        'filter_type': filter_type,
        'query_date': query_date,
        'query_date_end': query_date_end
    }
    return render(request, 'production/finishedgoods_message_request_history.html', context)


def finished_goods_message_history_pdf(request):
    # Retrieve and validate the query date parameters
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date_end else None

    data = Finished_Goods_Request.objects.all()
    # Apply date filter only if both parsed dates are valid
    if parsed_date and parsed_date_end:
        data = data.annotate(date_only=Cast('date', DateField())).filter(
            date_only__range=(parsed_date, parsed_date_end)  # Filter within the date range
        ).order_by('-date')
    elif parsed_date:
        data = data.annotate(date_only=Cast('date', DateField())).filter(
            date_only=parsed_date  # Filter within the date range
        ).order_by('-date')
    # Retrieve the first entry for dynamic filename generation
    view = data.first()
    filename = f"Production_Message_Requests_{query_date}_to_{query_date_end}.pdf" if view else "Production_Message_Requests.pdf"


    # Prepare context for rendering the PDF template
    context = {'view': data, 'name': view}

    # Load and render the template to HTML
    template = get_template('production/production_message_pdf.html')
    html = template.render(context)

    # Create the HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Generate the PDF
    pdf_status = pisa.CreatePDF(html, dest=response)

    if pdf_status.err:
        return HttpResponse("Error generating PDF", status=400)

    return response


def finished_goods_message_history_excel(request):
    # Retrieve and validate the query date parameters
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date_end else None

    # Fetch all finished good requests
    finished_good = Finished_Goods_Request.objects.all()

    # Apply date filter only if both parsed dates are valid
    if parsed_date and parsed_date_end:
        finished_good = finished_good.annotate(date_only=Cast('date', DateField())).filter(
            date_only__range=(parsed_date, parsed_date_end)  # Filter by date range
        ).order_by('-date')
    elif parsed_date:
        finished_good = finished_good.annotate(date_only=Cast('date', DateField())).filter(date_only=parsed_date).order_by('-date')

    # Get the category name for the filename
    filename = f"Production_Message_Requests_{query_date}_to_{query_date_end}.xlsx" if query_date and query_date_end else "Production_Message_Requests.xlsx"

    # Create an Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Production Message Request"

    # Write the title row at the top
    title = f"Production_Message_Requests_{query_date}_to_{query_date_end}" if query_date and query_date_end else "Production_Message_Requests"
    ws.append([title])  # Title row
    ws.append([])  # Blank row for line space
    ws.append([])  # Second blank row for line space

    # Write the header row (these are your "bold" headers in Excel)
    headers = ['Category', 'Stock', 'Date', 'Status']
    ws.append(headers)

    # Set header row font to bold
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)  # Header is in row 4 after title rows
        cell.font = header_font

    # Iterate through the filtered data and write each row to the Excel file
    for i in finished_good:
        ws.append([
            i.category.category_name,
            i.stock,
            i.date.strftime('%Y-%m-%d') if i.date else 'N/A',
            i.status
        ])

    # Prepare the response as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save the workbook to the response
    wb.save(response)

    return response


def shift_group(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    return render(request, 'production/shift_group1.html',{'allocated_menus': allocated_menus})


def shift_plan(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    return render(request,'production/shift_plan1.html', {'allocated_menus': allocated_menus})


def production_plan(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    return render(request, 'production/production_plan.html', {'allocated_menus': allocated_menus})


def shift_details(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    return render(request, 'production/shift_details.html', {'allocated_menus': allocated_menus})


def production_history(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    return render(request, 'production/production_history.html', {'allocated_menus': allocated_menus})


def shift_plan_history(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    return render(request, 'production/shift_plan_history.html', {'allocated_menus': allocated_menus})


def production_plan_history(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    return render(request, 'production/production_plan_history.html', {'allocated_menus': allocated_menus})

def post_production_history(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    return render(request, 'production/post_production_history.html', {'allocated_menus': allocated_menus})


def attendance_history(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    return render(request, 'production/attendance_history.html', {'allocated_menus': allocated_menus})


def shift_group_history(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    return render(request, 'production/shift_group_history.html', {'allocated_menus': allocated_menus})


def group_history(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    return render(request, 'production/group_history.html', {'allocated_menus': allocated_menus})


def production_inbox(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    return render(request, 'production/production_inbox.html', {'allocated_menus': allocated_menus})