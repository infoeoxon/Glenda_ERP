from django.db.models import F
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import get_template
from django.urls import reverse
from openpyxl.workbook import Workbook
from xhtml2pdf import pisa

from inventory_app.models import RawMaterialsStock
from purchase_app.models import RawMaterials
from rd_app.forms import RD_approvalForm
from Glenda_App.models import Menu
from rd_app.models import RD_stock,RD_Unapproved_Stock

from register_app.models import CustomUser, MenuPermissions
from django.core.paginator import Paginator
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.db.models.functions import Cast
from django.db.models import DateField
from xhtml2pdf import pisa
from openpyxl.styles import Font, Alignment
# Create your views here.

def raw_material_stock_approve(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)


    raw_materials_stock = RawMaterialsStock.objects.all().order_by('-date','-time')

    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parse_date_start = parse_date(query_date) if query_date else None
    parse_date_end = parse_date(query_date_end) if query_date_end else None
    is_filtered = False

    # Apply date range filtering if both start and end dates are provided
    if parse_date_start and parse_date_end:
        raw_materials_stock = raw_materials_stock.annotate(date_only=Cast('date', DateField())).filter(
            date_only__range=[parse_date_start, parse_date_end])
        is_filtered = True
    elif parse_date_start:
        raw_materials_stock = raw_materials_stock.annotate(date_only=Cast('date', DateField())).filter(date_only=parse_date_start)
        is_filtered = True

    # Pagination logic
    paginator = Paginator(raw_materials_stock, 5)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'raw_materials_stock':page_obj,
        'allocated_menus': allocated_menus,
        'query_date': query_date,
        'query_date_end': query_date_end,
        'is_filtered': is_filtered
    }

    return render(request,'rd/raw_material_stock_approval_list.html', context)


def raw_material_stock_approve_pdf(request):
    # Retrieve and validate the query date parameter
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date else None

    # Retrieve filtered or unfiltered data based on query date
    data = RawMaterialsStock.objects.all().order_by('-date','-time')
    if parsed_date and parsed_date_end:
        data = data.annotate(date_only=Cast('date', DateField())).filter(date_only__range=[parsed_date,parsed_date_end])
    elif parsed_date:
        data = data.annotate(date_only=Cast('date', DateField())).filter(date_only=parsed_date)

    # Retrieve the first entry for dynamic filename generation
    view = data.first()
    filename = f"Requests from Inventory.pdf" if view else "Inventory Request History.pdf"

    # Prepare context for rendering PDF
    context = {'view': data, 'name': view}

    # Load and render the template to HTML
    template = get_template('rd/pdf_raw_material_stock_approval_list.html')
    html = template.render(context)

    # Create the HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Generate the PDF
    pdf_status = pisa.CreatePDF(html, dest=response)

    return response

def raw_material_stock_approve_excel(request):
    # Retrieve and validate the query date parameters
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date_end else None

    # Retrieve filtered or unfiltered items based on the query date
    items = RawMaterialsStock.objects.all().order_by('-date','-time')
    if parsed_date and parsed_date_end:
        items = items.annotate(date_only=Cast('date', DateField())).filter(date_only__range=(parsed_date, parsed_date_end))
    elif parsed_date:
        items = items.annotate(date_only=Cast('date', DateField())).filter(date_only=parsed_date)

    # Generate filename based on query date parameters
    filename = f"Requests from Inventory_{query_date}to{query_date_end}.xlsx" if query_date and query_date_end else "Requests from Inventory.xlsx"

    # Create an Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "REQUESTS FROM INVENTORY"

    # Write the title row with bold formatting
    title = f"REQUESTS FROM INVENTORY {query_date} to {query_date_end}" if query_date and query_date_end else "REQUESTS FROM INVENTORY"
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)  # Set title font to bold and increase the size

    # Merge cells across columns for the title to center it
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    # Blank rows for spacing
    ws.append([])
    ws.append([])

    # Write the header row
    headers = ['REQUEST', 'STATUS', 'REQUESTED DATE', 'REQUESTED TIME', 'REMARKS']
    ws.append(headers)

    # Set header row font to bold
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)  # Header is in row 4 after title rows
        cell.font = header_font

    # Iterate through the filtered data and write each row to the Excel file
    for item in items:
        # Convert date to naive format and format as string for Excel compatibility
        date_value = item.date.strftime('%d-%m-%Y') if item.date else 'N/A'
        time_value = item.time.strftime('%H:%M:%S') if item.time else 'N/A'
        ws.append([
            item.raw_materials.category.category_name if item.raw_materials.category else 'N/A',
            item.status,
            date_value,
            time_value,
            item.remarks
        ])

    # Prepare the response as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save the workbook to the response
    wb.save(response)

    return response


def raw_material_stock_approve_review(request, id):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Your existing setup code
    request_data = get_object_or_404(RawMaterialsStock, id=id)
    page = request.GET.get('page', 1)

    if request.method == 'POST' and 'approve' in request.POST:
        try:
            approved_stock = int(request.POST.get('approved_stock', 0))
            reason = request.POST.get('response', '').strip()
        except ValueError:
            error_message = "Invalid number for approved stock."
            return render(request, 'rd/raw_material_stock_approval_view.html', {
                'data': request_data,
                'allocated_menus': allocated_menus,
                'page': page,
                'error_message': error_message,
            })

        # Validation checks
        if approved_stock > request_data.stock:
            error_message = "Approved stock cannot exceed the arrived stock quantity."
            return render(request, 'rd/raw_material_stock_approval_view.html', {
                'data': request_data,
                'allocated_menus': allocated_menus,
                'page': page,
                'error_message': error_message,
            })

        if approved_stock < request_data.stock and not reason:
            error_message = "Please provide a reason if not all stocks are approved."
            return render(request, 'rd/raw_material_stock_approval_view.html', {
                'data': request_data,
                'allocated_menus': allocated_menus,
                'page': page,
                'error_message': error_message,
            })

        # Update status and response fields without altering date and time
        RawMaterialsStock.objects.filter(id=request_data.id).update(
            status='Reviewed', response=reason
        )

        # Calculate unapproved stock
        unapproved_stock = request_data.stock - approved_stock

        # Add a new entry to RD_stock with the current date and time
        RD_stock.objects.create(
            raw_materials=request_data,
            approved_stock=approved_stock,
            unapproved_stock=unapproved_stock,
            date=timezone.now().date(),
            time=timezone.now().time()
        )

        # Update the total stock in RawMaterials
        raw_material = request_data.raw_materials
        raw_material.total_stock += approved_stock
        raw_material.save()

        # Redirect with page parameter
        return redirect(f'{reverse("raw_material_stock_approve")}?page={page}')

    # Render initial page
    return render(request, 'rd/raw_material_stock_approval_view.html', {
        'data': request_data,
        'allocated_menus': allocated_menus,
        'page': page,
    })


def stock_approval_history(request,id):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Use get_object_or_404 to handle missing RD_stock entries
    rd_stock = get_object_or_404(RD_stock, raw_materials_id=id)

    context = {
        'allocated_menus': allocated_menus,
        'data':rd_stock
    }

    return render(request,'rd/history_raw_material_stock_approval_view.html',context)


def verification_list(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    return render(request,'rd/verification_list.html', {'allocated_menus': allocated_menus})


def quality_check_list(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    return render(request, 'rd/quality_check_list.html', {'allocated_menus': allocated_menus})


def rd_history(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    return render(request, 'rd/rd_history.html', {'allocated_menus': allocated_menus})


def quality_check_history(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    return render(request, 'rd/quality_check_history.html', {'allocated_menus': allocated_menus})


def verification_history(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    return render(request, 'rd/verification_history.html', {'allocated_menus': allocated_menus})