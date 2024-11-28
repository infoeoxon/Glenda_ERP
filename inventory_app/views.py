import csv
import io
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.template.loader import get_template
from Glenda_App.models import Menu
from django.db.models import F
from django.contrib import messages
from inventory_app.forms import Raw_materials_StockForm, Finished_Goods_StockForm, Finished_Goods_RequestForm, \
    Damaged_Goods_StockForm, Raw_materials_requestForm, Raw_material_allocate_Form, Finished_Goods_allocateForm
from inventory_app.models import RawMaterialsStock, Finished_Goods_Stock, Finished_Goods_Request, Damaged_Goods_Stock, \
    Raw_material_request, Raw_material_allocate, Finished_Goods_allocate
from production_app.models import water_Finished_Goods, water_Finished_goods_category, damaged_Goods, \
    Damaged_good_category
from register_app.models import department
from purchase_app.models import RawMaterials, RawMaterialCategory
from rd_app.models import RD_stock
from Glenda_App.models import Menu
from xhtml2pdf import pisa
from django.db.models import Q
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
from django.http import FileResponse
from reportlab.pdfgen import canvas
from django.db.models import Sum
from django.utils.dateparse import parse_date
from django.db.models import DateField
from django.db.models.functions import Cast
from django.utils.dateparse import parse_date
from django.core.paginator import Paginator
from register_app.models import CustomUser, MenuPermissions




# Create your views here.


def raw_materials_view(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)


    raw_materials = RawMaterials.objects.prefetch_related('stocks').all()
    categories = RawMaterialCategory.objects.all()

    # Pagination logic
    paginator = Paginator(raw_materials, 5)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Calculate total stock for each raw material
    total_stocks = {
        material.id: material.stocks.aggregate(total=Sum('stock'))['total'] or 0
        for material in raw_materials
    }

    return render(request, 'inventory/view_raw_materials.html',{'view': raw_materials, 'allocated_menus': allocated_menus, 'total_stocks': total_stocks, 'categories': categories})


def finishedgoods_stock_view(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)


    finished_goods = water_Finished_Goods.objects.prefetch_related('stocks').all()
    categories = water_Finished_goods_category.objects.all()

    # Pagination logic
    paginator = Paginator(finished_goods, 5)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Calculate total stock for each raw material
    total_stocks = {
        material.id: material.stocks.aggregate(total=Sum('stock'))['total'] or 0
        for material in finished_goods
    }

    return render(request, 'inventory/view_finished_goods.html',
                  {'view': finished_goods, 'total_stocks': total_stocks, 'categories': categories, 'allocated_menus': allocated_menus, 'finished_goods': page_obj})


def finishedgoods_stock_history(request, id):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Get initial queryset for finished goods based on the provided ID
    finished_good = Finished_Goods_Stock.objects.filter(finished_goods_id=id).order_by('-date', '-time')

    # Get filter type and query date from the request
    filter_type = request.GET.get('filter')
    query_date = request.GET.get('query')  # Expected format: YYYY-MM-DD
    query_date_end = request.GET.get('query_end')  # Expected format: YYYY-MM-DD

    is_filtered = False

    # Apply filtering if filter type is 'date' and query_date is provided
    if filter_type == 'date' and query_date and query_date_end:
        finished_good = finished_good.filter(date__range=[query_date, query_date_end])
        is_filtered = True  # Flag to indicate that a filter has been applied

    # Pagination logic
    paginator = Paginator(finished_good, 5)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Calculate total stock for filtered results
    total_stock = sum(item.stock for item in finished_good)

    # Prepare context for rendering template
    context = {
        'data': page_obj,
        'allocated_menus': allocated_menus,
        'id': id,
        'filter_type': filter_type,
        'query_date': query_date,
        'is_filtered': is_filtered,
        'total_stock': total_stock,
        'query_date_end': query_date_end
    }

    return render(request, 'inventory/view_finished_goods_history.html', context)




def damagedgoods_stock_view(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    damaged_goods = damaged_Goods.objects.prefetch_related('stocks').all().order_by('-date', '-time')
    categories = Damaged_good_category.objects.all()

    # Calculate total stock for each raw material
    total_stocks = {
        material.id: material.stocks.aggregate(total=Sum('stock'))['total'] or 0
        for material in damaged_goods
    }

    # Pagination logic
    paginator = Paginator(damaged_goods, 5)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'inventory/view_damaged_goods.html',
                  {'view': damaged_goods, 'total_stocks': total_stocks, 'categories': categories, 'allocated_menus': allocated_menus, 'damaged_goods': page_obj})



def finished_goods_stock_history_pdf(request, id):
    # Retrieve and validate the query date parameters
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date_end else None

    # Apply date filter only if both parsed dates are valid
    if parsed_date and parsed_date_end:
        data = Finished_Goods_Stock.objects.filter(
            finished_goods_id=id,
            date__isnull=False,
            date__range=(parsed_date, parsed_date_end)  # Filter within the date range
        ).order_by('-date', '-time')
    else:
        data = Finished_Goods_Stock.objects.filter(finished_goods_id=id).order_by('-date', '-time')

    # Retrieve the first entry for dynamic filename generation
    view = data.first()
    filename = f"{view.finished_goods.category.category_name}_production_report.pdf" if view else "production_report.pdf"

    # Prepare context for rendering the PDF template
    context = {'view': data, 'name': view}

    # Load and render the template to HTML
    template = get_template('inventory/finishedgoods_stock_pdf.html')
    html = template.render(context)

    # Create the HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Generate the PDF
    pdf_status = pisa.CreatePDF(html, dest=response)

    if pdf_status.err:
        return HttpResponse("Error generating PDF", status=400)

    return response


def finishedgoods_stock_history_excel(request, id):
    # Retrieve and validate the query date parameters
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date_end else None

    # Apply date filter only if both parsed dates are valid
    if parsed_date and parsed_date_end:
        finished_good = Finished_Goods_Stock.objects.filter(
            finished_goods_id=id,
            date__isnull=False
        ).order_by('-date', '-time').filter(
            date__range=([parsed_date, parsed_date_end])  # Filter by date range
        )
    else:
        finished_good = Finished_Goods_Stock.objects.filter(finished_goods_id=id).order_by('-date', '-time')

    # Get the category name for the filename
    data = finished_good.first()  # Use first entry to get category info
    filename = f"{data.finished_goods.category.category_name}_inventory_report.xlsx" if data else "inventory_report.xlsx"

    # Create an Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Report"

    # Write the title row at the top
    title = f"{data.finished_goods.category.category_name} Inventory Report" if data else "Inventory Report"
    ws.append([title])  # Title row
    ws.append([])  # Blank row for line space
    ws.append([])  # Second blank row for line space

    # Write the header row (these are your "bold" headers in Excel)
    headers = ['Category', 'Name', 'Size', 'Stock', 'Date']
    ws.append(headers)

    # Set header row font to bold
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)  # Header is in row 4 after title rows
        cell.font = header_font

    # Iterate through the filtered data and write each row to the Excel file
    for i in finished_good:
        ws.append([
            i.finished_goods.category.category_name,
            i.finished_goods.name,
            i.finished_goods.size,
            i.stock,
            i.date.strftime('%Y-%m-%d') if i.date else 'N/A'  # Formatting date as needed
        ])

    # Prepare the response as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save the workbook to the response
    wb.save(response)

    return response



def raw_materials_stock_history(request, id):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Retrieve all RawMaterialsStock instances with the given raw_materials_id
    raw_materials_stock_qs = RawMaterialsStock.objects.filter(raw_materials_id=id)

    # Check if any raw materials stock is found
    if not raw_materials_stock_qs.exists():
        return render(request, '404.html')

    # Filter RD_stock entries associated with the retrieved RawMaterialsStock instances
    raw_materials = RD_stock.objects.filter(raw_materials__in=raw_materials_stock_qs).select_related(
        'raw_materials', 'raw_materials__raw_materials', 'raw_materials__raw_materials__category'
    )

    # Date-based filtering
    filter_type = request.GET.get('filter')
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parse_date_start = parse_date(query_date) if query_date else None
    parse_date_end = parse_date(query_date_end) if query_date_end else None
    is_filtered = False

    # Apply date range filtering if both start and end dates are provided
    if parse_date_start and parse_date_end:
        raw_materials = raw_materials.annotate(date_only=Cast('date', DateField())).filter(date_only__range=[parse_date_start, parse_date_end])
        is_filtered = True
    elif parse_date_start:
        raw_materials = raw_materials.annotate(date_only=Cast('date', DateField())).filter(date_only=parse_date_start)
        is_filtered = True

    # Calculate total stock for filtered results
    total_stock = sum(item.approved_stock for item in raw_materials)

    # Pagination logic
    paginator = Paginator(raw_materials, 5)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Pass the filtered results and query_date to the template
    context = {
        'data': page_obj,
        'allocated_menus': allocated_menus,
        'id': id,
        'filter_type': filter_type,
        'query_date': query_date,
        'query_date_end': query_date_end,
        'is_filtered': is_filtered,
        'total_stock':total_stock
    }

    return render(request, 'inventory/view_raw_materials_history.html', context)


def raw_materials_stock_pdf(request, id):
    # Retrieve and validate the query date parameters
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date_end else None

    # Retrieve all related RawMaterialsStock instances with the given raw_materials_id
    raw_materials_stock_qs = RawMaterialsStock.objects.filter(raw_materials_id=id)

    # Filter RD_stock entries associated with the retrieved RawMaterialsStock instances
    raw_materials = RD_stock.objects.filter(
        raw_materials__in=raw_materials_stock_qs
    ).select_related(
        'raw_materials', 'raw_materials__raw_materials', 'raw_materials__raw_materials__category'
    )

    # Apply date filter only if both parsed dates are valid
    if parsed_date and parsed_date_end:
        raw_materials = raw_materials.filter(date__range=(parsed_date, parsed_date_end))
    elif parsed_date:
        raw_materials = raw_materials.filter(date=parsed_date)

    # Get the name for the filename using the first entry in RD_stock
    view = raw_materials.first()  # Use the first entry to get the name for the report
    filename = f"{view.raw_materials.raw_materials.name}_raw_material_report.pdf" if view else "raw_material_report.pdf"

    # Prepare context for rendering PDF
    context = {'view': raw_materials, 'name': view}

    # Load and render the template
    template = get_template('inventory/raw_material_stock_pdf.html')
    html = template.render(context)

    # Create the HttpResponse object with the appropriate PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Generate the PDF
    pdf_status = pisa.CreatePDF(html, dest=response)

    # Check for any error in PDF generation and handle it
    if pdf_status.err:
        return HttpResponse("Error generating PDF", status=500)

    return response


def raw_materials_stock_excel(request, id):
    # Retrieve and validate the query date parameter
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date_end else None

    # Retrieve all related RawMaterialsStock instances with the given raw_materials_id
    raw_materials_stock_qs = RawMaterialsStock.objects.filter(raw_materials_id=id)

    # Filter RD_stock entries associated with the retrieved RawMaterialsStock instances
    raw_materials = RD_stock.objects.filter(
        raw_materials__in=raw_materials_stock_qs
    ).select_related(
        'raw_materials', 'raw_materials__raw_materials', 'raw_materials__raw_materials__category'
    )

    # Apply date filter only if both parsed dates are valid
    if parsed_date and parsed_date_end:
        raw_materials = raw_materials.filter(date__range=(parsed_date, parsed_date_end))
    elif parsed_date:
        raw_materials = raw_materials.filter(date=parsed_date)

    # Get the raw material name for the filename
    data = raw_materials.first()  # Use the first entry to get material info
    filename = f"{data.raw_materials.raw_materials.name}_raw_material_report.xlsx" if data else "raw_material_report.xlsx"

    # Create an Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Material Stock History"

    # Write the headline row that spans all columns
    ws.merge_cells('A1:E1')
    headline = f"{data.raw_materials.raw_materials.name} Raw Material Stock History" if data else "Raw Material Stock History"
    ws['A1'] = headline

    # Set font properties and center the headline
    headline_font = Font(bold=True, size=14)
    ws['A1'].font = headline_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # Write the header row
    headers = ['Category', 'Name', 'Size', 'Stock', 'Date']
    ws.append([])  # Blank row for spacing
    ws.append(headers)

    # Make the header row bold
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)  # Header is in row 3
        cell.font = header_font

    # Iterate through the data and write each row to the Excel file
    for item in raw_materials:
        ws.append([
            item.raw_materials.raw_materials.category.category_name if item.raw_materials.raw_materials.category else 'N/A',  # Category name
            item.raw_materials.raw_materials.name,  # Name of the raw material
            item.raw_materials.raw_materials.size,  # Size of the raw material
            item.approved_stock,  # Stock amount
            item.date.strftime('%Y-%m-%d') if item.date else 'N/A'  # Formatting date
        ])

    # Prepare the response as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save the workbook to the response
    wb.save(response)

    return response



# Function to render Damaged Goods Stock History with optional date filtering
def damagedgoods_stock_history(request, id):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Get initial queryset for damaged goods based on the provided ID
    damaged_goods = Damaged_Goods_Stock.objects.filter(damaged_id=id).order_by('-date', '-time')

    # Get filter type and query date from the request
    filter_type = request.GET.get('filter')
    query_date = request.GET.get('query')  # Expected format: YYYY-MM-DD
    query_date_end = request.GET.get('query_end')
    is_filtered = False

    # Apply filtering if filter type is 'date' and query_date is provided
    if filter_type == 'date' and query_date and query_date_end:
        damaged_goods = damaged_goods.filter(date__range=[query_date, query_date_end])
        is_filtered = True  # Flag to indicate that a filter has been applied

    # Pagination logic
    paginator = Paginator(damaged_goods, 5)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Calculate total stock for filtered results
    total_stock = sum(item.stock for item in damaged_goods)

    # Prepare context for rendering template
    context = {
        'id': id,
        'filter_type': filter_type,
        'query_date': query_date,
        'query_date_end': query_date_end,
        'is_filtered': is_filtered,  # Pass flag to template
        'allocated_menus': allocated_menus,
        'data': page_obj,
        'total_stock': total_stock
    }

    return render(request, 'inventory/damaged_stock_history.html', context)


def generate_pdf(request, id):
    # Retrieve and validate the query date parameters
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date_end else None

    # Apply date filter only if both parsed dates are valid
    if parsed_date and parsed_date_end:
        data = Damaged_Goods_Stock.objects.filter(
            damaged_id=id,
            date__isnull=False,
            date__range=(parsed_date, parsed_date_end)  # Filter within the date range
        ).order_by('-date', '-time')
    else:
        data = Damaged_Goods_Stock.objects.filter(damaged_id=id).order_by('-date', '-time')

    # Retrieve the first entry for dynamic filename generation
    view = data.first()
    filename = f"{view.damaged.category.category_name}_production_report.pdf" if view else "production_report.pdf"

    # Prepare context for rendering the PDF template
    context = {'view': data, 'name': view}

    # Load and render the template to HTML
    template = get_template('inventory/damaged_stock_pdf.html')
    html = template.render(context)

    # Create the HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Generate the PDF
    pdf_status = pisa.CreatePDF(html, dest=response)

    if pdf_status.err:
        return HttpResponse("Error generating PDF", status=400)

    return response


def damaged_stock_history_excel(request, id):
    # Retrieve and validate the query date parameters
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date_end else None

    # Apply date filter only if both parsed dates are valid
    if parsed_date and parsed_date_end:
        damaged_good = Damaged_Goods_Stock.objects.filter(
            damaged_id=id,
            date__isnull=False
        ).order_by('-date', '-time').filter(
            date__range=([parsed_date, parsed_date_end])  # Filter by date range
        )
    else:
        damaged_good = Damaged_Goods_Stock.objects.filter(damaged_id=id).order_by('-date', '-time')

    # Get the category name for the filename
    data = damaged_good.first()  # Use first entry to get category info
    filename = f"{data.damaged.category.category_name}_damaged_stock_report.xlsx" if data else "damaged_stock_report.xlsx"

    # Create an Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Damaged Stock Report"

    # Write the title row at the top
    title = f"{data.damaged.category.category_name} Damaged Stock Report" if data else "Damaged Stock Report"
    ws.append([title])  # Title row
    ws.append([])  # Blank row for line space
    ws.append([])  # Second blank row for line space

    # Write the header row (these are your "bold" headers in Excel)
    headers = ['Category', 'Name', 'Description', 'Stock', 'Date']
    ws.append(headers)

    # Set header row font to bold
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)  # Header is in row 4 after title rows
        cell.font = header_font

    # Iterate through the filtered data and write each row to the Excel file
    for i in damaged_good:
        ws.append([
            i.damaged.category.category_name,
            i.damaged.name,
            i.damaged.description,
            i.stock,
            i.date.strftime('%Y-%m-%d') if i.date else 'N/A'  # Formatting date as needed
        ])

    # Prepare the response as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save the workbook to the response
    wb.save(response)

    return response

def damaged_goods_export_and_search(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Get search query parameters
    search_name = request.GET.get('name', '').strip()
    search_category = request.GET.get('category', '').strip()

    # Start by showing all damaged goods if no search parameters are provided
    if not search_name and not search_category:
        damaged_goods_list = damaged_Goods.objects.all().order_by('-date', '-time')
    else:
        # If search filters are provided, apply them
        filters = Q()
        if search_name:
            filters &= Q(name__icontains=search_name)
        if search_category and search_category.isdigit():
            filters &= Q(category_id=search_category)

        # Apply the filters to the queryset
        damaged_goods_list = damaged_Goods.objects.filter(filters).order_by('-date', '-time')

        # If no results match the search, return an empty queryset
        if not damaged_goods_list.exists():
            damaged_goods_list = damaged_Goods.objects.none()

    # Check if the request is for exporting to Excel
    if 'export' in request.GET:
        # Create a Workbook and set up the worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = 'Damaged Goods List'

        # Merge and style the headline
        ws.merge_cells('A1:D1')
        ws['A1'] = 'Damaged Goods List'
        headline_font = Font(bold=True, size=14)
        ws['A1'].font = headline_font
        ws['A1'].alignment = Alignment(horizontal='center')

        # Write the header row
        headers = ['Category', 'Name', 'Description', 'Total Stock']
        ws.append(headers)

        # Make header row bold
        header_font = Font(bold=True)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.font = header_font

        # Write data rows
        for damaged_good in damaged_goods_list:
            ws.append([
                damaged_good.category.category_name if damaged_good.category else 'N/A',
                damaged_good.name,
                damaged_good.description,
                damaged_good.total_stock
            ])

        # Create an HTTP response for the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="damaged_goods_list.xlsx"'

        # Save the workbook to the response
        wb.save(response)

        return response

    # Otherwise, render the filtered search results to the template
    categories = Damaged_good_category.objects.all()
    context = {
        'view': damaged_goods_list,
        'categories': categories,
        'search_name': search_name,
        'search_category': search_category,
        'allocated_menus': allocated_menus
    }

    return render(request, 'inventory/view_damaged_goods.html', context)

def finishedgoods_search_and_export(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Get search query parameters
    search_name = request.GET.get('name', '').strip()  # Get name from the query string
    search_category = request.GET.get('category', '').strip()  # Get category from the query string

    # Start by showing all finished goods if no search parameters are provided
    if not search_name and not search_category:
        finished_goods_list = water_Finished_Goods.objects.all().order_by('-date', '-time')  # Show all data initially
    else:
        # If search filters are provided, apply them
        filters = Q()  # Initialize an empty filter

        if search_name:
            filters &= Q(name__icontains=search_name)  # Filter by name if provided

        if search_category and search_category.isdigit():
            filters &= Q(category_id=search_category)  # Filter by category if provided

        # Apply the filters to the queryset
        finished_goods_list = water_Finished_Goods.objects.filter(filters).order_by('-date', '-time')

        # If no results match the search, return an empty queryset
        if not finished_goods_list.exists():
            finished_goods_list = water_Finished_Goods.objects.none()  # Return an empty queryset

    # Check if the request is for exporting to Excel
    if 'export' in request.GET:
        # Create a Workbook and set up the worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = 'Finished Goods List'

        # Merge and style the headline
        ws.merge_cells('A1:D1')
        ws['A1'] = 'Finished Goods List'
        headline_font = Font(bold=True, size=14)
        ws['A1'].font = headline_font
        ws['A1'].alignment = Alignment(horizontal='center')

        # Write the header row
        headers = ['Category', 'Name', 'Size', 'Total Stock']
        ws.append(headers)

        # Make header row bold
        header_font = Font(bold=True)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.font = header_font

        # Write data rows
        for finished_good in finished_goods_list:
            ws.append([
                finished_good.category.category_name if finished_good.category else 'N/A',
                finished_good.name,
                finished_good.size,
                finished_good.total_stock
            ])

        # Create an HTTP response for the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="finished_goods_list.xlsx"'

        # Save the workbook to the response
        wb.save(response)

        return response

    # Otherwise, render the filtered search results to the template
    categories = water_Finished_goods_category.objects.all()
    context = {
        'view': finished_goods_list,
        'categories': categories,
        'search_name': search_name,
        'search_category': search_category,
        'allocated_menus': allocated_menus
    }

    return render(request, 'inventory/view_finished_goods.html', context)


def finishedgoods_request_messages_list(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)


    data = Finished_Goods_Request.objects.all().order_by('-date')

    # Get filter type and query date from the request
    filter_type = request.GET.get('filter')
    query_date = request.GET.get('query')  # Expected format: YYYY-MM-DD
    query_date_end = request.GET.get('query_end')
    is_filtered = False

    # Apply filtering if filter type is 'date' and query_date is provided
    if filter_type == 'date' and query_date and query_date_end:
        data = data.annotate(date_only=Cast('date', DateField())).filter(date_only__range=[query_date, query_date_end])
        is_filtered = True  # Flag to indicate that a filter has been applied

    elif query_date:
        data = data.annotate(date_only=Cast('date', DateField())).filter(date_only=query_date)
        is_filtered = True  # Flag to indicate that a filter has been applied

    # Pagination logic
    paginator = Paginator(data, 5)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'data': page_obj,
        'allocated_menus': allocated_menus,
        'is_filtered': is_filtered,
        'filter_type': filter_type,
        'query_date': query_date,
        'query_date_end': query_date_end
    }
    return render(request, 'inventory/finishedgoods_request_message_list.html', context)


def finishedgoods_message_request_list_pdf(request):
    # Retrieve and validate the query date parameters
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date_end else None

    data = Finished_Goods_Request.objects.all()
    # Apply date filter only if both parsed dates are valid
    if parsed_date and parsed_date_end:
        data = data.annotate(date_only=Cast('date', DateField())).filter(
            date_only__range=(parsed_date, parsed_date_end)  # Filter within the date range
        ).order_by('-date')
    elif parsed_date:
        data = data.annotate(date_only=Cast('date', DateField())).filter(
            date_only=parsed_date  # Filter within the date range
        ).order_by('-date')
    # Retrieve the first entry for dynamic filename generation
    view = data.first()
    filename = f"Message_Requests_{query_date}_to_{query_date_end}.pdf" if view else "Inventory_Message_Requests.pdf"

    # Prepare context for rendering the PDF template
    context = {'view': data, 'name': view}

    # Load and render the template to HTML
    template = get_template('production/production_message_pdf.html')
    html = template.render(context)

    # Create the HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Generate the PDF
    pdf_status = pisa.CreatePDF(html, dest=response)

    if pdf_status.err:
        return HttpResponse("Error generating PDF", status=400)

    return response

def finishedgoods_message_request_list_excel(request):
    # Retrieve and validate the query date parameters
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date_end else None

    # Fetch all finished good requests
    finished_good = Finished_Goods_Request.objects.all().order_by('-date')

    # Apply date filter only if both parsed dates are valid
    if parsed_date and parsed_date_end:
        finished_good = finished_good.annotate(date_only=Cast('date', DateField())).filter(
            date_only__range=(parsed_date, parsed_date_end)  # Filter by date range
        ).order_by('-date')
    elif parsed_date:
        finished_good = finished_good.annotate(date_only=Cast('date', DateField())).filter(date_only=parsed_date).order_by('-date')

    # Get the category name for the filename
    filename = f"Production_Message_Requests_{query_date}_to_{query_date_end}.xlsx" if query_date and query_date_end else "Production_Message_Requests.xlsx"

    # Create an Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Production Message Request"

    # Write the title row at the top
    title = f"Message_Requests_{query_date}_to_{query_date_end}" if query_date and query_date_end else "Inventory_Message_Requests"
    ws.append([title])  # Title row
    ws.append([])  # Blank row for line space
    ws.append([])  # Second blank row for line space

    # Write the header row (these are your "bold" headers in Excel)
    headers = ['Category', 'Stock', 'Remarks', 'Date', 'Status']
    ws.append(headers)

    # Set header row font to bold
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)  # Header is in row 4 after title rows
        cell.font = header_font

    # Iterate through the filtered data and write each row to the Excel file
    for i in finished_good:
        ws.append([
            i.category.category_name,
            i.stock,
            i.remarks,
            i.date.strftime('%Y-%m-%d') if i.date else 'N/A',
            i.status
        ])

    # Prepare the response as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save the workbook to the response
    wb.save(response)

    return response

def finishedgoods_message_request(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    dept = department.objects.all()
    category = water_Finished_goods_category.objects.all()
    name = water_Finished_Goods.objects.all()
    view = Finished_Goods_Request.objects.all()

    if request.method == 'POST':
        form = Finished_Goods_RequestForm(request.POST)
        print(request.POST)  # Debug: Print the form data
        if form.is_valid():
            print("Form is valid")  # Debug: Form is valid
            form_entry = form.save(commit=False)
            # Ensure foreign key is assigned correctly
            form_entry.name = form.cleaned_data['name']
            form_entry.status = 'Pending'
            form_entry.save()
            return redirect('finishedgoods_request_messages_list')
        else:
            print(form.errors)  # Debug: Print form errors
    else:
        form = Finished_Goods_RequestForm()

    return render(request, 'inventory/finishedgoods_message_request.html', {
        'form': form,
        'department': dept,
        'category': category,
        'name': name,
        'view': view,
        'allocated_menus': allocated_menus
    })


def raw_material_search_and_export(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Get search query parameters
    search_name = request.GET.get('name', '').strip()
    search_category = request.GET.get('category', '').strip()

    # Start by showing all raw materials if no search parameters are provided
    if not search_name and not search_category:
        raw_material_list = RawMaterials.objects.all()
    else:
        filters = Q()
        if search_name:
            filters &= Q(name__icontains=search_name)
        if search_category and search_category.isdigit():
            filters &= Q(category_id=search_category)
        raw_material_list = RawMaterials.objects.filter(filters)
        if not raw_material_list.exists():
            raw_material_list = RawMaterials.objects.none()

    # Check if the request is for exporting to Excel
    if 'export' in request.GET:
        # Create a Workbook and set up the worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = 'Raw Materials List'

        # Merge and style the headline
        ws.merge_cells('A1:D1')
        ws['A1'] = 'Raw Materials List'
        headline_font = Font(bold=True, size=14)
        ws['A1'].font = headline_font
        ws['A1'].alignment = Alignment(horizontal='center')

        # Write the header row
        headers = ['Category', 'Name', 'Size', 'Total Stock']
        ws.append(headers)

        # Make header row bold
        header_font = Font(bold=True)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.font = header_font

        # Write data rows
        for raw_material in raw_material_list:
            ws.append([
                raw_material.category.category_name if raw_material.category else 'N/A',
                raw_material.name,
                raw_material.size,
                raw_material.total_stock
            ])

        # Create an HTTP response for the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="raw_materials_list.xlsx"'

        # Save the workbook to the response
        wb.save(response)

        return response

    # Otherwise, render the filtered search results to the template
    categories = RawMaterialCategory.objects.all()
    context = {
        'view': raw_material_list,
        'categories': categories,
        'search_name': search_name,
        'search_category': search_category,
        'allocated_menus': allocated_menus
    }

    return render(request, 'inventory/view_raw_materials.html', context)


def raw_material_request_list(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    message = Raw_material_request.objects.all().order_by('-date')


    # Pagination logic
    paginator = Paginator(message, 5)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'inventory/raw_material_request_list.html', {'allocated_menus': allocated_menus, 'message': page_obj})



def raw_material_message_request(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    dept = department.objects.all()
    category = RawMaterialCategory.objects.all()
    name = RawMaterials.objects.all()
    view = Raw_material_request.objects.all()

    if request.method == 'POST':
        form = Raw_materials_requestForm(request.POST)
        if form.is_valid():
            form_entry = form.save(commit=False)
            form_entry.name = form.cleaned_data['name']
            form_entry.status = 'Pending'
            form_entry.save()

        return redirect('raw_material_request_list')  # Redirect to the list view
    else:
        form = Raw_materials_requestForm()

    return render(request, 'inventory/raw_material_message_request.html', {
        'form': form,
        'department': dept,
        'category': category,
        'name': name,
        'view': view,
        'allocated_menus': allocated_menus
    })


def raw_material_allocate(request, id):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    raw_material = get_object_or_404(RawMaterials, id=id)

    if request.method == 'POST':
        form = Raw_material_allocate_Form(request.POST)
        if form.is_valid():
            stock_allocate = form.save(commit=False)
            allocated_stock = stock_allocate.stock

            # Update the total stock for the raw material by deducting allocated stock
            if allocated_stock > raw_material.total_stock:
                messages.error(request, "Not enough stock available")
                return render(request, 'inventory/raw_material_allocate.html', {
                    'allocated_menus': allocated_menus,
                    'raw_material': raw_material,
                    'form': form,
                })

            stock_allocate.raw_material = raw_material  # Set the raw material
            stock_allocate.save()  # Save the new stock entry

            raw_material.total_stock = F('total_stock') - allocated_stock
            raw_material.save(update_fields=["total_stock"])  # Save the updated total stock

            return redirect('Raw_materials_view')  # Redirect to the list view

    else:
        form = Raw_material_allocate_Form()

    return render(request, 'inventory/raw_material_allocate.html', {

        'raw_material': raw_material,
        'form': form,
        'message': messages,
        'allocated_menus': allocated_menus
    })



def raw_material_allocate_history(request, id):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Retrieve and validate the query date parameter
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date_start = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date_end else None
    is_filtered = False

    # Get initial queryset for allocated stock based on provided ID
    allocated_stock = Raw_material_allocate.objects.filter(raw_material_id=id).order_by('-date')

    # Apply date filtering if parsed_date is valid
    if parsed_date_start and parsed_date_end:
        allocated_stock = allocated_stock.annotate(date_only=Cast('date', DateField())).filter(date_only__range=[parsed_date_start,parsed_date_end])
        is_filtered = True

    elif parsed_date_start:
        allocated_stock = allocated_stock.annotate(date_only=Cast('date', DateField())).filter(
            date_only=parsed_date_start
        )
        is_filtered = True

    # Calculate total stock for filtered results
    total_stock = sum(item.stock for item in allocated_stock)

    # Pagination logic
    paginator = Paginator(allocated_stock, 5)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Prepare context for rendering the template
    context = {
        'allocated_menus': allocated_menus,
        'allocated_stock': page_obj,
        'id': id,
        'query_date': query_date,
        'query_date_end': query_date_end,
        'is_filtered': is_filtered,
        'total_stock':total_stock
    }

    return render(request, 'inventory/raw_material_allocate_history.html', context)

def raw_material_allocate_excel(request, id):
    # Retrieve and validate the query date parameter
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date_end else None

    # Retrieve filtered or unfiltered items based on the query date
    items = Raw_material_allocate.objects.filter(raw_material_id=id)
    if parsed_date and parsed_date_end:
        items = items.annotate(date_only=Cast('date', DateField())).filter(date_only__range=[parsed_date,parsed_date_end])
    elif parsed_date:
        items = items.annotate(date_only = Cast('date',DateField())).filter(date_only=parsed_date)

    # Retrieve the first entry for dynamic filename generation
    data = items.first()
    filename = f"{data.raw_material.category.category_name}_allocate_stock_history.xlsx" if data else "allocate_stock_history.xlsx"

    # Create a Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Material Allocate History"

    # Write a headline row and center it
    ws.merge_cells('A1:E1')
    ws[
        'A1'] = f"{data.raw_material.category.category_name} Allocate Stock History" if data else "Allocate Stock History"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Write the header row
    headers = ['Category', 'Name', 'Size', 'Stock', 'Date']
    ws.append([])
    ws.append(headers)

    # Apply bold font to header row
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = header_font

    # Write data rows
    for item in items:
        ws.append([
            item.raw_material.category.category_name if item.raw_material.category else 'N/A',
            item.raw_material.name,
            item.raw_material.size,
            item.stock,
            item.date.strftime('%Y-%m-%d') if item.date else 'N/A'
        ])

    # Create an HTTP response with the appropriate Excel headers
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response


def raw_materials_stock_allocate_pdf(request, id):
    # Retrieve and validate the query date parameter
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date else None

    # Retrieve filtered or unfiltered data based on query date
    data = Raw_material_allocate.objects.filter(raw_material_id=id)
    if parsed_date and parsed_date_end:
        data = data.annotate(date_only=Cast('date', DateField())).filter(date_only__range=[parsed_date,parsed_date_end])
    elif parsed_date:
        data = data.annotate(date_only=Cast('date', DateField())).filter(date_only=parsed_date)

    # Retrieve the first entry for dynamic filename generation
    view = data.first()
    filename = f"{view.raw_material.name}_allocate_history.pdf" if view else "allocate_history.pdf"

    # Prepare context for rendering PDF
    context = {'view': data, 'name': view}

    # Load and render the template to HTML
    template = get_template('inventory/raw_material_allocate_stock_pdf.html')
    html = template.render(context)

    # Create the HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Generate the PDF
    pdf_status = pisa.CreatePDF(html, dest=response)

    return response



def finished_goods_allocate(request, id):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)


    finished_goods = get_object_or_404(water_Finished_Goods, id=id)

    if request.method == 'POST':
        form = Finished_Goods_allocateForm(request.POST)
        if form.is_valid():
            stock_entry = form.save(commit=False)
            new_stock = stock_entry.stock  # Stock to allocate

            # Check if new stock allocation exceeds total stock
            if new_stock > finished_goods.total_stock:
                messages.error(request, "Allocation exceeds available total stock.")
                return render(request, 'inventory/finished_goods_allocate.html', {
                    'form': form,
                    'allocated_menus': allocated_menus,
                    'finished_goods': finished_goods
                })

            # Proceed with allocation if within available stock
            stock_entry.finished_good = finished_goods
            stock_entry.save()

            # Decrease the total stock by the allocated amount
            finished_goods.total_stock = F('total_stock') - new_stock
            finished_goods.save(update_fields=["total_stock"])

            return redirect('finishedgoods_stock_view')  # Redirect to the list view
    else:
        form = Finished_Goods_StockForm()

    return render(request, 'inventory/finished_goods_allocate.html', {
        'form': form,
        'finished_goods': finished_goods,
        'allocated_menus': allocated_menus
    })


def finished_goods_allocate_history(request, id):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Get initial queryset for allocated stock based on provided ID
    allocated_stock = Finished_Goods_allocate.objects.filter(finished_good_id=id).order_by('-date')

    # Get filter type and query date from the request
    filter_type = request.GET.get('filter')
    query_date = request.GET.get('query')  # Expected format: YYYY-MM-DD
    query_date_end = request.GET.get('query_end')  # Expected format: YYYY-MM-DD

    is_filtered = False

    # Apply filtering if filter type is 'date' and query_date is provided
    if filter_type == 'date' and query_date and query_date_end:
        allocated_stock = allocated_stock.annotate(date_only=Cast('date', DateField())).filter(date_only__range=[query_date, query_date_end])
        is_filtered = True  # Flag to indicate that a filter has been applied

    # Pagination logic
    paginator = Paginator(allocated_stock, 5)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Calculate total stock for filtered results
    total_stock = sum(item.stock for item in allocated_stock)

    # Pagination logic
    paginator = Paginator(allocated_stock, 5)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Prepare context for rendering the template
    context = {

        'id': id,
        'filter_type': filter_type,
        'query_date': query_date,
        'query_date_end': query_date_end,
        'is_filtered': is_filtered,
        'allocated_menus': allocated_menus,
        'allocated_stock': page_obj,
        'total_stock': total_stock
    }

    return render(request, 'inventory/finished_goods_allocate_history.html', context)


def finished_goods_allocate_excel(request, id):
    # Retrieve and validate the query date parameters
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date_end else None

    # Get initial queryset for allocated stock based on provided ID
    items = Finished_Goods_allocate.objects.filter(finished_good_id=id).order_by('-date')

    # Handle filtering if parsed_date is valid
    if parsed_date and parsed_date_end:
        items = items.annotate(date_only=Cast('date', DateField())).filter(date_only__range=[parsed_date, parsed_date_end]).order_by('-date')

    # Get the category name for the filename
    data = items.first()
    filename = f"{data.finished_good.category.category_name}_allocate_stock_history.xlsx" if data else "allocate_stock_history.xlsx"

    # Create a Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Allocate Stock History"

    # Write a headline row that spans all columns
    ws.merge_cells('A1:E1')
    ws['A1'] = f"{data.finished_good.category.category_name} Allocate Stock History" if data else "Allocate Stock History"

    # Set font properties and center the headline
    headline_font = Font(bold=True, size=14)
    ws['A1'].font = headline_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # Write the header row
    headers = ['Category', 'Name', 'Size', 'Allocated Stock', 'Remarks','Date']
    ws.append(headers)

    # Make header row bold
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)  # Changed to row 2
        cell.font = header_font

    # Write data rows
    for item in items:
        ws.append([
            item.finished_good.category.category_name if item.finished_good.category else 'N/A',
            item.finished_good.name,
            item.finished_good.size,
            item.stock,
            item.remarks,
            item.date.strftime('%Y-%m-%d') if item.date else 'N/A'
        ])

    # Create an HTTP response with the appropriate Excel header
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save the workbook to the response
    wb.save(response)

    return response


def finished_goods_stock_allocate_pdf(request, id):
    # Retrieve and validate the query date parameters
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date_end else None

    # Handle filtering if parsed_date is valid
    if parsed_date and parsed_date_end:
        data = Finished_Goods_allocate.objects.filter(finished_good_id=id, date__isnull=False)
        data = data.annotate(date_only=Cast('date', DateField())).filter(date_only__range=(parsed_date, parsed_date_end)).order_by('-date')
    else:
        data = Finished_Goods_allocate.objects.filter(finished_good_id=id).order_by('-date')

    # Get the name for the filename using the first entry
    view = data.first()
    filename = f"{view.finished_good.name}_allocate_history.pdf" if view else "allocate_history.pdf"

    # Prepare context for rendering PDF
    context = {'view': data, 'name': view}

    # Load and render the template
    template = get_template('inventory/pdf_finishedgoods_allocate_history.html')
    html = template.render(context)

    # Create the HttpResponse object with the appropriate PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Generate the PDF
    pdf_status = pisa.CreatePDF(html, dest=response)

    if pdf_status.err:
        return HttpResponse('Error generating PDF', status=500)

    return response



def inbox(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    return render(request,'inventory/inbox.html', {'allocated_menus': allocated_menus})


def stock_verification(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    return render(request,'inventory/stock_verification.html', {'allocated_menus': allocated_menus})


def demo_inventory_overview(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    return render(request,'inventory/demo_inventory_overview.html',{'allocated_menus': allocated_menus})

def demo_request_list(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    return render(request, 'inventory/demo_request_list.html', {'allocated_menus': allocated_menus})


def demo_all_inventory_history(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    return render(request, 'inventory/demo_all_inventory_history.html', {'allocated_menus': allocated_menus})


def demo_request_list_history(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    return render(request, 'inventory/demo_request_list_history.html', {'allocated_menus': allocated_menus})


def demo_stock_history(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    return render(request, 'inventory/demo_stock_history.html', {'allocated_menus': allocated_menus})


def demo_arrived_stock_verification_history(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated
    # submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    return render(request, 'inventory/demo_arrived_stock_verification_history.html', {'allocated_menus': allocated_menus})


