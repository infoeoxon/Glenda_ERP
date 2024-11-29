import json
from datetime import datetime, timedelta, date
import calendar
from django.shortcuts import render,redirect,get_object_or_404,reverse
from register_app.models import CustomUser, MenuPermissions
from hr_app.models import EmployeeDetails, RequestLeave, Payroll, Attendance, Remarks, Resignation, AttendanceReport
from Glenda_App.models import Menu
from .forms import EmployeeDetailsForm,LeaveRequestForm,PayrollForm,ResignationForm,BasicForm
from django.db.models import Q
from django.http import HttpResponse
from django.contrib import messages
from django.template.loader import render_to_string, get_template
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import csv
from decimal import Decimal
from django.utils import timezone
from django.http import HttpResponse
from django.db.models import Count
from django.db.models import Sum, Count
from django.core.management.base import BaseCommand
from django.utils.timezone import now


@login_required
def employee_list_and_verify(request):
    user = request.user

    # Fetch allocated menus and sub-menus
    user_permissions = MenuPermissions.objects.filter(user=user).select_related('menu', 'sub_menu')
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Handle verification actions if a specific employee is targeted
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        action = request.POST.get('action')
        if employee_id and action:
            employee = get_object_or_404(EmployeeDetails, user_id=employee_id)
            if action == 'approve':
                employee.verified = True
                verified = employee.verified
                return redirect(reverse('add_employee_details', args=[employee.id]))
            elif action == 'reject':
                employee.status = 'rejected'
                employee.save()
                messages.error(request, "Employee details rejected.")
            return redirect('employee_list_and_verify')

    # Fetch users with verified employee details
    users_with_details = CustomUser.objects.filter(
        is_staff=True,
        is_superuser=False,
        employeedetails__status='Verified'
    ).distinct()

    # Add `details_added` flag to users
    for user in users_with_details:
        user.details_added = EmployeeDetails.objects.filter(user=user).exists()

    # Paginate the results
    paginator = Paginator(users_with_details, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'allocated_menus': allocated_menus,
    }

    return render(request, 'hr/Employee_list.html', context)


@login_required
def employee_details_modal(request):
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        employee = get_object_or_404(EmployeeDetails, user_id=employee_id)

        # Prepare data for modal
        data = {
            'name': employee.user.name,
            'email': employee.user.email,
            'phone_number': employee.user.phone_number,
            'joining_date': employee.joining_date,
            'address': f"{employee.street}, {employee.district}, {employee.state}, {employee.country}",
            'blood_group': employee.employee_blood_groups,
            'date_of_birth':employee.date_of_birth,
            'id_proof_type':employee.id_proof_type,
            'id_proof_type_number':employee.id_proof_type_number,
            'id_proof_type_validity':employee.validity
        }

        return JsonResponse(data)

@login_required
def verify_employee_by_senior(request):
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        employee = get_object_or_404(EmployeeDetails, user_id=employee_id)

        # Update employee status to verified
        employee.status = 'Approved'
        employee.save()

        return JsonResponse({'success': True})

    return JsonResponse({'success': False}, status=400)

@login_required
def verify_employee(request):
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        employee = get_object_or_404(EmployeeDetails, user_id=employee_id)

        # Update employee status to verified
        employee.status = 'verified'
        employee.save()

        return JsonResponse({'success': True})

    return JsonResponse({'success': False}, status=400)


@login_required
def AddDetails(request):

    use = request.user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    if request.method == 'POST':
        form = EmployeeDetailsForm(request.POST, request.FILES)

        if form.is_valid():
            employee = form.save(commit=False)  # Do not save yet
            employee.user = use
            employee.status = 'pending_verification'
            employee.save()
            messages.success(request, "Details submitted successfully for verification.")
            return redirect('view_my_profile')

        else:
            # Form contains errors
            messages.error(request, "Please correct the errors below.")

    else:
        form = EmployeeDetailsForm()

    return render(request, 'hr/add_employee_details.html', {'form': form, 'allocated_menus': allocated_menus})

def view_employee_profile(request,id):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    view = get_object_or_404(EmployeeDetails,user_id=id)

    return render(request,'hr/view_employee_profile.html',{'view':view,'allocated_menus':allocated_menus})

def update_employee_details(request,id):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    view = get_object_or_404(EmployeeDetails, user_id=id)

    if request.method == 'POST':
        form = EmployeeDetailsForm(request.POST,instance=view)

        if form.is_valid():
            form.save()
            return redirect('Employee_list')

    else:
        form = EmployeeDetailsForm(request.POST, instance=view)
    return render(request, 'hr/update_employee_details.html', {'view':view, 'allocated_menus': allocated_menus,'form':form})


@login_required
def resign_employee_details(request, id):
    user = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=user).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Fetch the employee details or return a 404 error if not found
    dtl = EmployeeDetails.objects.filter(user_id=id)

    if request.method == "POST":
        form = ResignationForm(request.POST)

        if form.is_valid():
            resignation = form.save(commit=False)
            resignation.employee = dtl  # Link resignation to employee details
            resignation.status = Resignation.APPROVED_BY_SENIOR  # Set status appropriately
            resignation.save()  # Save resignation record

            messages.success(request, f"Employee {dtl.user.name}'s resignation has been processed.")

    else:
        form = ResignationForm()  # Initialize an empty form for GET requests

    return render(request, 'hr/my_profile.html', {
        'form': form,
        'allocated_menus': allocated_menus,
        'dtl': dtl,
    })


def employee_search_and_export(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Start with an empty query to show all staff members by default
    employee_list = EmployeeDetails.objects.select_related('user').filter(
        user__is_staff=True, user__is_superuser=False
    )

    search_query = request.GET.get('search_query', '').strip()
    export_format = request.GET.get('export', '')

    if search_query:
        # Create a Q object to filter based on the search query
        filters = Q(user__is_staff=True, user__is_superuser=False)  # Base filter to restrict staff members
        if search_query.isdigit():
            filters &= Q(user__phone_number__icontains=search_query) | Q(pincode__icontains=search_query)
        else:
            filters &= Q(user__name__icontains=search_query) | Q(state__icontains=search_query)

        # Apply the filters
        employee_list = EmployeeDetails.objects.select_related('user').filter(filters)

    # Check for export format (either 'excel' or 'pdf')
    if export_format == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Employee List'
        ws.merge_cells('A1:D1')
        ws['A1'] = 'Employee List'
        headline_font = Font(bold=True, size=14)
        ws['A1'].font = headline_font
        ws['A1'].alignment = Alignment(horizontal='center')

        headers = ['Name', 'Email', 'Phone Number', 'State', 'Pincode']
        ws.append(headers)

        header_font = Font(bold=True)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.font = header_font

        for employee in employee_list:
            ws.append([
                employee.user.name,
                employee.user.email,
                employee.user.phone_number,
                employee.state,
                employee.pincode,
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="employee_list.xlsx"'

        wb.save(response)
        return response

    elif export_format == 'pdf':
        # PDF export
        template = get_template('hr/employee_details_pdf.html')  # A separate template for PDF
        context = {'employees': employee_list, 'allocated_menus': allocated_menus}
        html = template.render(context)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="employee_list.pdf"'

        pdf_status = pisa.CreatePDF(html, dest=response)

        if pdf_status.err:
            return HttpResponse('Error generating PDF', status=500)

        return response

    # Add `details_added` flag to each employee in the list
    for employee in employee_list:
        employee.details_added = (
            employee.basic and employee.pf_no and employee.employee_esi_no
        )

    context = {
        'employees': employee_list,  # Filtered employees list
        'allocated_menus': allocated_menus,
    }

    return render(request, 'hr/Employee_list.html', context)



def view_leave_list(request):
    # Fetch all employees
    employees = EmployeeDetails.objects.all()

    # Prepare a list to hold employee leave requests
    leave_requests = []

    for employee in employees:
        # Get all leave requests for each employee
        requests = RequestLeave.objects.filter(employee=employee,verification_status='approved')
        leave_requests.append({
            'employee': employee,
            'requests': requests
        })

    # Apply pagination to leave_requests list
    paginator = Paginator(leave_requests, 10)  # Show 10 employees per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Pass the paginated leave_requests and menus to the context
    context = {
        'allocated_menus': allocated_menus,
        'page_obj': page_obj  # Include the page object in the context
    }

    return render(request, 'hr/leave_list.html', context)


@csrf_exempt
def approve_leave_request(request):
    if request.method == 'POST':
        user_id = request.POST.get('employee_user_id')
        try:
            leave_request = RequestLeave.objects.get(id=user_id)
            leave_request.approval_status = 'APPROVED'  # Update status to 'approved'
            leave_request.save()

            start_date = leave_request.start_date
            end_date = leave_request.end_date

            date = start_date
            while date <= end_date:
                attendance, created = Attendance.objects.get_or_create(
                    employee=leave_request.employee,
                    date=date,
                )
                attendance.status = 'Leave'
                attendance.save()
                date += timedelta(days=1)

            return JsonResponse({'success': True})
        except RequestLeave.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Leave request not found'})

@csrf_exempt
def reject_leave_request(request):
    if request.method == 'POST':
        user_id = request.POST.get('employee_user_id')
        rejection_reason = request.POST.get('rejection_reason')
        try:
            leave_request = RequestLeave.objects.get(id=user_id)
            leave_request.approval_status = 'REJECTED'  # Update status to 'rejected'
            leave_request.rejection_reason = rejection_reason
            leave_request.save()
            return JsonResponse({'success': True})
        except RequestLeave.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Leave request not found'})



def my_leave_request(request):
    user = request.user
    leave_requests = RequestLeave.objects.filter(employee__user=user).order_by('-start_date')  # Order by date
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Pagination logic
    paginator = Paginator(leave_requests, 5)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'hr/my_leave_request.html', {
        'user': user,
        'leave_requests': page_obj,
        'allocated_menus': allocated_menus,
    })


def leave_request_form(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    user = request.user
    emp=EmployeeDetails.objects.get(user_id=user)
    if request.method == 'POST':
        form = LeaveRequestForm(request.POST)
        if form.is_valid():
            leave_request = form.save(commit=False)
            # Set the employee based on the logged-in user
            leave_request.employee_id =emp.id
            leave_request.save()
            return redirect('my_leave_request')  # Redirect after saving
    else:
        form = LeaveRequestForm()

    return render(request, 'hr/leave_request_form.html', {'form': form, 'allocated_menus': allocated_menus})


def search_approved_leave_requests(request):
    # Fetch all employees
    employees = EmployeeDetails.objects.all()

    # Prepare a list to hold employee leave requests
    leave_requests = []

    # Get search query from request
    search_query = request.GET.get('search_query', '').strip()

    for employee in employees:
        # Get all approved leave requests for each employee
        requests = RequestLeave.objects.filter(employee=employee, approval_status='Approved')

        print(requests)

        # If there is a search query, filter by employee name
        if search_query and search_query.lower() not in employee.user.name.lower():
            continue  # Skip this employee if the name does not match the search query

        leave_requests.append({
            'employee': employee,
            'requests': requests  # This will be a QuerySet of leave requests
        })

    # Apply pagination to leave_requests list
    paginator = Paginator(leave_requests, 10)  # Show 10 employees per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Pass the paginated leave_requests and menus to the context
    context = {
        'allocated_menus': allocated_menus,
        'page_obj': page_obj,  # Include the page object in the context
        'search_query': search_query,  # Include search query for form repopulation
    }

    return render(request, 'hr/leave_list.html', context)

def employee_detail(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Get initial queryset for finished goods based on the provided ID
    Employee = EmployeeDetails.objects.all()

    # Get filter type and query date from the request
    filter_type = request.GET.get('filter')
    query_date = request.GET.get('query')  # Expected format: YYYY-MM-DD
    is_filtered = False

    # Apply filtering if filter type is 'date' and query_date is provided
    if filter_type == 'date' and query_date:
        employee_list = Employee.filter(date=query_date)
        is_filtered = True  # Flag to indicate that a filter has been applied

    # Prepare context for rendering template
    context = {
        'data': Employee,
        'allocated_menus': allocated_menus,
        'filter_type': filter_type,
        'query_date': query_date,
        'is_filtered': is_filtered  # Pass flag to template
    }

    return render(request, 'hr/employee_detail.html', context)

def leave_history(request, id):
    user = request.user  # Get the current user

    # Fetch user's permissions
    user_permissions = MenuPermissions.objects.filter(user=user).select_related('menu', 'sub_menu')

    # Organize menus and submenus in a dictionary
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Fetch leave requests for the specified employee
    leave_requests = RequestLeave.objects.filter(employee__user_id=id).order_by('-start_date')

    leave_data = []
    for leave in leave_requests:
        # Generate a list of dates from start_date to end_date
        date_range = []
        current_date = leave.start_date
        while current_date <= leave.end_date:
            date_range.append(current_date)
            current_date += timedelta(days=1)

        # Generate a list of weekdays corresponding to the date range
        weekdays_with_dates = [
            {'weekday': calendar.day_name[date.weekday()], 'date': date} for date in date_range
        ]

        leave_data.append({
            'date_range': date_range,
            'leave': leave,
            'weekdays_with_dates': weekdays_with_dates,
        })

    context = {
        'allocated_menus': allocated_menus,
        'leave_data': leave_data,
    }

    return render(request, 'hr/leave_history.html', context)


def create_payroll(request, id):

    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)


    # Retrieve the employee instance based on the given user ID
    employee = get_object_or_404(EmployeeDetails, user_id=id)
    basic_salary = employee.basic or Decimal('0.00')  # Ensure fallback to 0 if basic salary is null

    if request.method == 'POST':
        form = PayrollForm(request.POST)
        if form.is_valid():

            payroll_date = form.cleaned_data['pay_date']

            # Check if a payroll entry already exists for the same employee in the same month and year
            existing_payroll = Payroll.objects.filter(
                employee=employee,
                pay_date__year=payroll_date.year,
                pay_date__month=payroll_date.month
            ).exists()

            if existing_payroll:
                messages.warning(request,f"A payroll entry for {employee.user.name} already exists for {payroll_date.strftime('%B %Y')}.")
                return render(request, 'hr/payroll_form.html',{'form': form, 'employee': employee, 'basic_salary': basic_salary})

            payroll = form.save(commit=False)
            payroll.employee = employee  # Set the employee field for Payroll

            # Calculate gross earnings
            gross_earnings = (
                    basic_salary +
                    form.cleaned_data.get('hra', Decimal('0.00')) +
                    form.cleaned_data.get('da', Decimal('0.00')) +
                    form.cleaned_data.get('ot', Decimal('0.00')) +
                    form.cleaned_data.get('bonuses', Decimal('0.00'))+
                    form.cleaned_data.get('medical',Decimal('0.00'))+
                    form.cleaned_data.get('insurance',Decimal('0.00'))+
                    form.cleaned_data.get('ta', Decimal('0.00'))+
                    form.cleaned_data.get('food_allowance', Decimal('0.00'))
            )

            # Calculate total deductions
            total_deductions = (
                    form.cleaned_data.get('canteen_expense', Decimal('0.00')) +
                    form.cleaned_data.get('pf', Decimal('0.00')) +
                    form.cleaned_data.get('esi', Decimal('0.00')) +
                    form.cleaned_data.get('leave_plus', Decimal('0.00'))
            )

            # Calculate net pay and set it on the payroll instance
            payroll.net_pay = gross_earnings - total_deductions

            # Save the payroll instance with the calculated net pay
            payroll.save()

            # Redirect to the desired page after successful save
            return redirect('Employee_list')
    else:
        # Provide the basic salary to the form if needed
        form = PayrollForm()

    # Render the form along with basic salary and employee information
    return render(request, 'hr/payroll_form.html', {'form': form, 'employee': employee, 'basic_salary': basic_salary,'allocated_menus':allocated_menus})

def payroll_summary(request,id):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    payroll = Payroll.objects.filter(employee__user_id=id).order_by('-pay_date')

    # Get filter type and query date from the request
    filter_type = request.GET.get('filter')
    query_date = request.GET.get('query')  # Expected format: YYYY-MM-DD
    is_filtered = False

    # Apply filtering if filter type is 'date' and query_date is provided
    if filter_type == 'date' and query_date:
        finished_good = payroll.filter(date=query_date)
        is_filtered = True  # Flag to indicate that a filter has been applied

    # Prepare context for rendering template
    context = {
        'data': payroll,
        'allocated_menus': allocated_menus,
        'filter_type': filter_type,
        'query_date': query_date,
        'is_filtered': is_filtered  # Pass flag to template
    }

    return render(request, 'hr/payroll_summary.html', context)


def payroll_by_month(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Get current month and year
    current_month = timezone.now().month
    current_year = timezone.now().year

    # Get month and year from GET parameters, defaulting to current values
    selected_month = request.GET.get('month', current_month)
    selected_year = request.GET.get('year', current_year)

    # Convert selected_month and selected_year to integers
    try:
        selected_month = int(selected_month)
        selected_year = int(selected_year)
    except ValueError:
        selected_month = current_month
        selected_year = current_year

    # Fetch payroll records for the selected month and year
    payroll_records = Payroll.objects.filter(pay_date__month=selected_month,
                                             pay_date__year=selected_year)

    if not payroll_records.exists():
        messages.warning(request, 'No payroll records found for this month.')

    total_net_pay = sum(record.net_pay for record in payroll_records)

    # Define months as tuples (month_number, month_name)
    months = [
        (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
        (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
        (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
    ]

    context = {
        'allocated_menus': allocated_menus,
        'payrolls': payroll_records,
        'total_net_pay': total_net_pay,
        'months': months,
        'selected_month': selected_month,
        'selected_year': selected_year,
    }

    return render(request, 'hr/salary_list_by_month.html', context)

def add_employee_details(request, id):
    employee = get_object_or_404(EmployeeDetails, user_id=id)
    if request.method == "POST":
        form = BasicForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            return redirect("employee_list_and_verify")  # Adjust to your URL name
        else:
            # If the form is not valid, render the modal with the errors
            return render(request, "hr/Employee_list.html", {"form": form, "employee": employee})
    else:
        form = BasicForm(instance=employee)
    return render(request, "hr/Employee_list.html", {"form": form, "employee": employee})

@login_required
def employee_request_list(request):
    user = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=user).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Fetch users with employee details that need verification
    users_with_details = CustomUser.objects.filter(
        is_staff=True,
        is_superuser=False,
        employeedetails__isnull=False
    ).distinct()

    # Add `details_added` flag to users
    for user in users_with_details:
        user.details_added = EmployeeDetails.objects.filter(user=user).exists()

    # Handle verification actions if a specific employee is targeted
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        action = request.POST.get('action')
        employee = get_object_or_404(EmployeeDetails, user_id=employee_id)

        if action == 'verify':
            employee.status = 'verified'
            employee.save()
            messages.success(request, f"Employee {employee.user.name}'s details verified successfully.")
            return redirect('employee_request_list')

        elif action == 'reject':
            employee.status = 'rejected'
            employee.save()
            messages.error(request, f"Employee {employee.user.name}'s details were rejected.")
            return redirect('employee_request_list')

    # Paginate the results
    paginator = Paginator(users_with_details, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Pass the filtered HR employee details and menus to the context
    context = {
        'page_obj': page_obj,
        'allocated_menus': allocated_menus,
    }

    return render(request, 'hr/Employee_request_list.html', context)

def view_my_profile(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Try to fetch EmployeeDetails object
    try:
        view = EmployeeDetails.objects.get(user=use)
    except EmployeeDetails.DoesNotExist:
        view = None  # If not found, set to None

    return render(request, 'hr/my_profile.html', {'view': view, 'allocated_menus': allocated_menus})

def verify_leave_list(request):
    # Fetch all employees
    employees = EmployeeDetails.objects.all()

    # Prepare a list to hold employee leave requests
    leave_requests = []

    for employee in employees:
        # Get all leave requests for each employee
        requests = RequestLeave.objects.filter(employee=employee)
        leave_requests.append({
            'employee': employee,
            'requests': requests  # This will be a QuerySet of leave requests
        })

    # Apply pagination to leave_requests list
    paginator = Paginator(leave_requests, 10)  # Show 10 employees per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Pass the paginated leave_requests and menus to the context
    context = {
        'allocated_menus': allocated_menus,
        'page_obj': page_obj  # Include the page object in the context
    }

    return render(request, 'hr/Leave_list_verification.html', context)

@csrf_exempt
def verify_leave_request(request):
    if request.method == 'POST':
        user_id = request.POST.get('employee_user_id')
        try:
            leave_request = RequestLeave.objects.get(id=user_id)
            leave_request.verification_status = 'APPROVED'  # Update status to 'approved'
            leave_request.save()
            return JsonResponse({'success': True})
        except RequestLeave.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Leave request not found'})

@csrf_exempt
def reject_verify_leave_request(request):
    if request.method == 'POST':
        user_id = request.POST.get('employee_user_id')
        rejection_reason = request.POST.get('rejection_reason')
        try:
            leave_request = RequestLeave.objects.get(id=user_id)
            leave_request.verification_status = 'REJECTED'  # Update status to 'rejected'
            leave_request.rejection_reason = rejection_reason
            leave_request.save()
            return JsonResponse({'success': True})
        except RequestLeave.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Leave request not found'})

def calculate_monthly_attendance(employee_id, month, year):
    # Get the start and end dates of the month
    start_date = datetime(year, month, 1)
    end_date = (datetime(year + (month // 12), (month % 12) + 1, 1) - timedelta(days=1)).date()

    # Query attendance records for the given month
    attendance = Attendance.objects.filter(
        employee_id=employee_id,
        date__range=[start_date, end_date]
    ).values('status').annotate(count=Count('status'))

    # Get the total days in the month
    _, total_days_in_month = calendar.monthrange(year, month)

    all_dates = [start_date + timedelta(days=i) for i in range(total_days_in_month)]

    # Filter weekdays (Monday to Friday)
    weekdays = [day for day in all_dates if day.weekday() < 5]

    absent_days = sum(record['count'] for record in attendance if record['status'] == 'Absent')
    leave_days = sum(record['count'] for record in attendance if record['status'] == 'Leave')

    # Calculate present days
    present_days = len(weekdays) - (absent_days + leave_days)

    # Build the attendance report
    report = {
        "Present": present_days,
        "Absent": absent_days,
        "Leave": leave_days,
    }

    return report


@login_required
def attendance_view(request):
    user = request.user
    today = now().date()

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=user).select_related('menu', 'sub_menu')
    allocated_menus = {perm.menu: [] for perm in user_permissions}

    for perm in user_permissions:
        allocated_menus[perm.menu].append(perm.sub_menu)

    filter_type = request.GET.get('filter', 'daily')
    query_date = request.GET.get('query')
    selected_month = request.GET.get('month')
    selected_year = request.GET.get('year')

    employees = EmployeeDetails.objects.all()

    # Initialize context
    context = {
        'allocated_menus': allocated_menus,
        'filter_type': filter_type,
        'employees': employees,
        'data': []  # Initialize data as an empty list
    }

    if filter_type == 'daily':
        filter_date = datetime.strptime(query_date, '%Y-%m-%d').date() if query_date else today
        attendance_records = Attendance.objects.filter(date=filter_date).select_related('employee')

        # Map attendance records by employee ID
        attendance_map = {record.employee.id: record for record in attendance_records}

        leave_dates = [record.date.strftime('%Y-%m-%d') for record in attendance_records if record.status == 'Leave']
        present_dates = [record.date.strftime('%Y-%m-%d') for record in attendance_records if record.status == 'Present']
        absent_dates = [record.date.strftime('%Y-%m-%d') for record in attendance_records if record.status == 'Absent']

        # Prepare data for each employee
        for employee in employees:
            record = attendance_map.get(employee.id)
            context['data'].append({
                'name': employee.user.name,
                'present_dates': json.dumps(present_dates),  # Ensure these are valid JSON strings
                'leave_dates': json.dumps(leave_dates),
                'absent_dates': json.dumps(absent_dates),
            })

        context['query_date'] = filter_date.strftime('%Y-%m-%d')

    elif filter_type == 'monthly':

      return render(request, 'hr/attendance_view.html', context)

    return render(request, 'hr/attendance_view.html', context)

def senior_resignation_approval(request, id):
    use = request.user
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    status = request.post.get('status')

    dtl = Resignation.objects.filter(employee_id=id,status='ACTIVE')

    if request.method == "POST":
        # Senior approves the resignation
        dtl.senior_approval = True
        dtl.resignation_status = Resignation.APPROVED_BY_HR  # Move to HR for final approval
        dtl.save()
        return redirect('hr_resignation_approval', id=dtl.id)  # Redirect to HR for approval

    return render(request, 'hr/senior_resignation_approval.html', {'dtl': dtl, 'allocated_menus': allocated_menus})

def hr_resignation_approval(request, id):
    use = request.user
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    dtl = get_object_or_404(EmployeeDetails, id=id)

    if request.method == "POST":
        # HR approves the resignation
        dtl.hr_approval = True
        dtl.resignation_status = Resignation.RESIGNED  # Mark the employee as resigned
        dtl.status = Resignation.RESIGNED  # Change employee status to resigned
        dtl.save()
        return redirect('employee_list')  # Redirect to the employee list page

    return render(request, 'hr/hr_resignation_approval.html', {'dtl': dtl, 'allocated_menus': allocated_menus})

def Resign_list_for_senior(request):
    use = request.user


    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')


    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    users_with_details = Resignation.objects.filter(
        status = 'Pending'
    ).distinct()

    for user in users_with_details:

        user.details_added = EmployeeDetails.objects.filter(user=user).exists()

    paginator = Paginator(users_with_details,5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)


    context = {
        'page_obj':page_obj,
        'allocated_menus': allocated_menus,

    }

    return render(request, 'hr/Resign_list_for_senior.html', context)

def Resign_list_for_hr(request):
    use = request.user


    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')


    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    users_with_details = Resignation.objects.filter(
        status = 'Approved by Senior'
    ).distinct()

    for user in users_with_details:

        user.details_added = EmployeeDetails.objects.filter(user=user).exists()

    paginator = Paginator(users_with_details,5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)


    context = {
        'page_obj':page_obj,
        'allocated_menus': allocated_menus,

    }

    return render(request, 'hr/Resign_list_for_hr.html', context)

@csrf_exempt
def mark_attendance(request):
    use = request.user
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        action = request.POST.get('action')  # 'check_in', 'check_out', or 'status_update'

        # Validate employee existence
        try:
            employee = EmployeeDetails.objects.get(pk=employee_id)
        except EmployeeDetails.DoesNotExist:
            return JsonResponse({'error': 'Employee not found'}, status=404)

        # Retrieve or create the attendance record for today
        attendance, created = Attendance.objects.get_or_create(employee=employee, date=date.today())

        if action in ['check_in', 'check_out']:
            # Prevent marking absent or leave after check-in/check-out
            if attendance.status in ['Absent', 'Leave']:
                return JsonResponse({'error': 'Cannot check in or out after marking absent or leave'}, status=400)

        if action == 'check_in':
            if attendance.check_in_time:
                return JsonResponse({'error': 'Check-in already recorded'}, status=400)
            attendance.check_in_time = datetime.now().time()
            attendance.status = 'Present'

        elif action == 'check_out':
            if attendance.check_out_time:
                return JsonResponse({'error': 'Check-out already recorded'}, status=400)
            attendance.check_out_time = datetime.now().time()

        elif action == 'status_update':
            new_status = request.POST.get('status')  # 'Present', 'Absent', or 'Leave'
            # Validate the status
            if new_status not in dict(Attendance._meta.get_field('status').choices):
                return JsonResponse({'error': 'Invalid status'}, status=400)
            attendance.status = new_status

        else:
            return JsonResponse({'error': 'Invalid action'}, status=400)

        attendance.save()
        return JsonResponse({'message': 'Attendance updated successfully'})

    # If not POST, render the attendance page
    employees = EmployeeDetails.objects.all()
    return render(request, 'hr/attendance.html', {'employees': employees,'allocated_menus':allocated_menus})



class Command(BaseCommand):
    help = "Generate attendance reports for the previous month"

    def handle(self, *args, **kwargs):
        today = datetime.today()
        first_day_of_current_month = datetime(today.year, today.month, 1)
        last_day_of_previous_month = first_day_of_current_month - timedelta(days=1)
        first_day_of_previous_month = datetime(last_day_of_previous_month.year, last_day_of_previous_month.month, 1)

        # Get all employees
        employees = Attendance.objects.values('employee').distinct()

        for emp in employees:
            employee = emp['employee']
            attendances = Attendance.objects.filter(
                employee_id=employee,
                date__range=(first_day_of_previous_month, last_day_of_previous_month)
            )

            total_days_worked = attendances.filter(status='Present').count()
            total_leaves = attendances.filter(status='Leave').count()
            total_absent = attendances.filter(status='Absent').count()
            total_hours_worked = sum(
                attendance.worked_hours for attendance in attendances if attendance.worked_hours
            )

            # Create or update the report
            AttendanceReport.objects.update_or_create(
                employee_id=employee,
                report_month=first_day_of_previous_month,
                defaults={
                    'total_days_worked': total_days_worked,
                    'total_leaves': total_leaves,
                    'total_absent': total_absent,
                    'total_hours_worked': total_hours_worked,
                },
            )

        self.stdout.write(self.style.SUCCESS("Attendance reports generated successfully."))

@csrf_exempt
def fetch_attendance(request):
    if request.method == 'GET':
        today = now().date()
        attendance_records = Attendance.objects.filter(date=today).select_related('employee')
        data = [
            {
                'employee_id': attendance.employee.id,
                'employee_name': attendance.employee.user.name,
                'status': attendance.status,
                'check_in_time': attendance.check_in_time.strftime('%H:%M:%S') if attendance.check_in_time else 'N/A',
                'check_out_time': attendance.check_out_time.strftime('%H:%M:%S') if attendance.check_out_time else 'N/A'
            }
            for attendance in attendance_records
        ]
        return JsonResponse({'attendance': data}, safe=False)
    return JsonResponse({'error': 'Invalid request method'}, status=400)

def attendance_reports(request):
    reports = AttendanceReport.objects.all().select_related('employee').order_by('-report_month')
    return render(request, 'hr/attendance_reports.html', {'reports': reports})