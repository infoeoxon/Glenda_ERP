from Glenda_App.models import Menu, SubMenu
from register_app.forms import CustomUserForm, CustomLoginForm, designation_Form, department_Form, \
    PermissionForm
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth import authenticate
from django.db import IntegrityError
from django.contrib.auth import login,logout
from register_app.models import CustomUser, designation, MenuPermissions, department
from django.core.paginator import Paginator
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment
from django.http import HttpResponse
from django.core.paginator import Paginator

def staff_home(request):
    user = request.user

    # Get the user's primary key ID
    user_id = user.id
    # Debug: Print user info
    print(user_id)
    # Get the Menu_permisions instance for the current user
    permissions = MenuPermissions.objects.filter(
        user_id=user_id).first()  # Get the first permission instance for the user

    # Debug: Print permissions

    if permissions:
        # Extract menu IDs from permissions
        menu_ids = permissions.menu_details.values_list('id', flat=True)
        # Filter menus based on permissions
        menus = Menu.objects.filter(id__in=menu_ids).prefetch_related('submenus')
    else:
        # If no permissions are found, return an empty queryset
        menus = Menu.objects.none()

    # Debug: Print filtered menus
    print(f"Filtered Menus: {menus}")
    return render(request,'register/home/index.html',{'menus':menus})



def register_view(request):



    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Define your custom order (e.g., by menu ID or another attribute)
    custom_order = ['Master', 'Purchase', 'Production','Inventory','Logistics','R & D','HR','Customer','Vendor','Accounts']  # replace with your actual menu IDs or names

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Sort menus based on custom order
    sorted_menus = sorted(
        allocated_menus.items(),
        key=lambda x: custom_order.index(x[0].title) if x[0].title in custom_order else float('inf')
    )
    sorted_allocated_menus = dict(sorted_menus)

    # Create a new dictionary for sorted menus
    import random

    # Generate a random number
    random_number = random.randint(1000, 9999999)

    # Concatenate "GA" with the random number
    result = f"GA{random_number}"

    print(result)
    if request.method == 'POST':
        form = CustomUserForm(request.POST,request.FILES)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_staff = True
            user.unique_id=result
            # Set is_staff to True
            user.save()  # Save the user instance

            # Log the user in after successful registration
            messages.success(request, "Registration successful!")
            return redirect('view_users')

        messages.error(request, "Please correct the errors below.")

    else:
        # Create an empty form for a GET request
        form = CustomUserForm()

    # Render the registration page with the form and menus
    return render(request, 'register/create_user.html', {'form': form, 'allocated_menus': sorted_allocated_menus})


def Edit_user(request,id):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    user = get_object_or_404(CustomUser, id=id)

    if request.method == 'POST':
        form = CustomUserForm(request.POST,request.FILES,instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, "User updated successfully!")
            return redirect('view_users')  # Redirect to a relevant page (e.g., user list or dashboard)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = CustomUserForm(instance=user)

    return render(request, 'register/update_users.html', {'form': form,'allocated_menus':allocated_menus})


def delete_user_view(request, user_id):
    # Retrieve the user object to be deleted
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)


    us = get_object_or_404(CustomUser, id=user_id)

    if request.method == 'POST':
        # If confirmed, delete the user
        us.delete()
        messages.success(request, "User deleted successfully!")
        return redirect('view_users')  # Redirect to a list of users or homepage

    # If request is GET, prompt confirmation
    return render(request, 'register/delete_user.html', {'us': us,'allocated_menus':allocated_menus})



# def login_view(request):
#     if request.method == 'POST':
#         form = CustomLoginForm(request.POST)
#         if form.is_valid():
#             email = form.cleaned_data['email']
#             password = form.cleaned_data['password']
#             user = authenticate(request, email=email, password=password)
#             if user is not None:
#                 login(request, user)
#                 # Redirect based on user role
#                 if user.is_superuser:
#                     return redirect('admin_home')  # Redirect to admin dashboard for superusers
#                 department_name = user.department.dept_Name if user.department else None
#                 designation_name = user.designation.user_type if user.designation else None
#
#                 # Redirect based on department and designation
#                 if designation_name == 'Executive' and department_name == 'Sales':
#                     return redirect('sales_home')
#                 elif designation_name == 'Manager' and department_name == 'Sales':
#                     return redirect('vender_home')
#                 elif designation_name == 'Assistant Manager' and department_name == 'Purchase':
#                     return redirect('manager_home')
#                 elif designation_name == 'Executive' and department_name == 'Logistics':
#                     return redirect('manager_home')
#                 else:
#                     return redirect('default_dashboard')  # Default page for other users
#     else:
#         form = CustomLoginForm()
#
#     return render(request, 'login.html', {'form': form})


def logout_view(request):
    logout(request)
    # Redirect to a specific page after logout (e.g., home page)
    return redirect(reverse('mainindex'))


def add_department(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    if request.method == 'POST':
        form = department_Form(request.POST)  # Use request.FILES for file upload
        if form.is_valid():
            form.save()
            return redirect('add_designation')  # Redirect to a list view or another page after saving
    else:
        form = department_Form()

    return render(request, 'register/create_department.html', {'form': form,'allocated_menus':allocated_menus})


def add_designation(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    if request.method == 'POST':
        form = designation_Form(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Designation added successfully!')
            return redirect('view_users')
    else:
        form = designation_Form()

    return render(request, 'register/create_designation.html', {'form': form,'allocated_menus':allocated_menus})


def view_users(request):
    form = CustomUserForm()

    sort_order = request.GET.get('sort', 'asc')  # Default sorting is ascending
    search_query = request.GET.get('search', '')  # Get search query
    selected_department = request.GET.get('department', None)  # Get selected department filter

    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Define your custom order (e.g., by menu ID or another attribute)
    custom_order = ['Master', 'Purchase', 'Production', 'Inventory', 'Logistics', 'R & D', 'HR', 'Customer', 'Vendor', 'Accounts']

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Sort menus based on custom order
    sorted_menus = sorted(
        allocated_menus.items(),
        key=lambda x: custom_order.index(x[0].title) if x[0].title in custom_order else float('inf')
    )

    # Create a new dictionary for sorted menus
    sorted_allocated_menus = dict(sorted_menus)

    # Start with a query for CustomUser, excluding superusers and non-staff users
    view = CustomUser.objects.filter(is_superuser=False, is_staff=True)

    # Apply department filter if selected
    if selected_department:
        view = view.filter(department__dept_Name=selected_department)

    # Apply search query if present
    if search_query:
        view = view.filter(
            Q(name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone_number__icontains=search_query)
        )

    # Apply sorting
    if sort_order == 'asc':
        view = view.order_by('name')  # Sort by name in ascending order
    elif sort_order == 'desc':
        view = view.order_by('-name')  # Sort by name in descending order

    # Pagination
    paginator = Paginator(view, 5)  # Show 5 users per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'form':form,
        'sort_order': sort_order,
        'selected_department': selected_department,
        'page_obj': page_obj,
        'allocated_menus': sorted_allocated_menus,
        'search_query': search_query,  # Pass search query to template for persistence
    }

    return render(request, 'register/view_users.html', context)

def delete_multiple_designations(request):
    if request.method == 'POST':
        selected_ids = request.POST.get('selected_ids', '').split(',')
        designation.objects.filter(id__in=selected_ids).delete()
        return redirect('view_designation')  # Redirect to the designation list page
    return HttpResponse(status=405)  # Method Not Allowed if not POST
def get_item(dictionary, key):
    """Get value from a dictionary using the key."""
    return dictionary.get(key)

def create_permission(request, id):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Define your custom order (e.g., by menu ID or another attribute)
    custom_order = ['Master', 'Purchase', 'Production', 'Inventory', 'Logistics', 'R & D', 'HR', 'Customer', 'Vendor',
                    'Accounts']  # replace with your actual menu IDs or names

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Sort menus based on custom order
    sorted_menus = sorted(
        allocated_menus.items(),
        key=lambda x: custom_order.index(x[0].title) if x[0].title in custom_order else float('inf')
    )

    # Create a new dictionary for sorted menus
    sorted_allocated_menus = dict(sorted_menus)

    if request.method == 'POST':
        form = PermissionForm(request.POST)
        if form.is_valid():
            # Handle selected menus
            menu_ids = request.POST.getlist('menu')  # List of selected menu IDs

            for menu_id in menu_ids:
                submenu_ids = request.POST.getlist(f'sub_menu_{menu_id}')  # Get submenus for the current menu

                for submenu_id in submenu_ids:
                    # Check if the permission already exists
                    existing_permission = MenuPermissions.objects.filter(
                        user_id=id,
                        menu_id=menu_id,
                        sub_menu_id=submenu_id
                    ).exists()

                    if not existing_permission:
                        try:
                            # Create a new MenuPermissions entry
                            MenuPermissions.objects.create(
                                user_id=id,
                                menu_id=menu_id,
                                sub_menu_id=submenu_id
                            )
                        except IntegrityError:
                            messages.warning(request,
                                             f"Error occurred while adding permission for Menu ID: {menu_id} and SubMenu ID: {submenu_id}.")
                    else:
                        messages.warning(request,
                                         f"Permission already exists for User ID: {id}, Menu ID: {menu_id}, and SubMenu ID: {submenu_id}.")

            messages.success(request, "Permissions saved successfully.")
            return redirect('admin_home')  # Redirect after saving

    else:
        form = PermissionForm()

    return render(request, 'register/add_permissions.html', {
        'form': form,
        'allocated_menus': sorted_allocated_menus,
    })



def get_submenus(request, menu_id):
    submenus = SubMenu.objects.filter(menu_id=menu_id).values('id', 'title')
    return JsonResponse(list(submenus), safe=False)


from django.http import JsonResponse

def load_designations(request):
    department_id = request.GET.get('department_id')
    designations = designation.objects.filter(dept_id=department_id).values('id', 'user_type')  # Adjust field names as necessary
    return JsonResponse(list(designations), safe=False)


def login_view(request):
    if request.method == 'POST':
        form = CustomLoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']
            user = authenticate(request, email=email, password=password)
            if user is not None:
                login(request, user)

                # Check if user is superuser and redirect to admin dashboard
                if user.is_superuser:
                    return redirect('management_index')

                # Redirect based on department name
                department_name = user.department.dept_Name if user.department else None
                if department_name == 'BOD':
                    return redirect('management_index')
                elif department_name == 'Admin':
                    return redirect('management_index')
                elif department_name == 'Sales':
                    return redirect('sales_index')
                elif department_name == 'Purchase':
                    return redirect('purchase_dashboard')
                elif department_name == 'Inventory':
                    return redirect('Inventory_index')
                elif department_name == 'Logistics':
                    return redirect('logistics_index')
                elif department_name == 'Production':
                    return redirect('production_index')
                elif department_name == 'R & D':
                    return redirect('rd_dashboard')
                elif department_name == 'HR':
                    return redirect('Hr_index')
                elif department_name == 'Accounts':
                    return redirect('accounts_index')
                else:
                    return redirect('admin')  # Default fallback

            else:
                form.add_error(None, "Invalid email or password.")
    else:
        form = CustomLoginForm()

    return render(request, 'login.html', {'form': form})


def logout_view(request):
    logout(request)
    # Redirect to a specific page after logout (e.g., home page)
    return redirect(reverse('admin'))

# def  user_search_export(request):
#     menus = Menu.objects.prefetch_related('submenus').all()

def  user_search(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)


    user_list = CustomUser.objects.all()  # Default to all vendors
    export_format = request.GET.get('export', '')

    if request.method == 'GET':
        search_query = request.GET.get('search_query', '').strip()

        if search_query:
            # Build filters
            filters = Q()

        if search_query.isdigit():
            filters &= Q(phone_number__icontains=search_query)  # Filter by user name

        else:
            filters &= Q(name__icontains=search_query)  # Filter by vendor phone number

        # Apply filters if any were provided
        if filters:
            user_list = CustomUser.objects.filter(filters)

        if export_format == 'excel':
            wb = Workbook()
            ws = wb.active
            ws.title = 'User List'
            ws.merge_cells('A1:D1')
            ws['A1'] = 'User List'
            headline_font = Font(bold=True, size=14)
            ws['A1'].font = headline_font
            ws['A1'].alignment = Alignment(horizontal='center')

            headers = ['Name', 'Email', 'Phone Number']
            ws.append(headers)

            header_font = Font(bold=True)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_num)
                cell.font = header_font

            for user in user_list:
                ws.append([user.name, user.email, user.phone_number])

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="employee_list.xlsx"'

            wb.save(response)
            return response


    context = {
        'page_obj': user_list,  # This will be used in the template
        'allocated_menus': allocated_menus,
    }

    return render(request, 'register/view_users.html', context)






def view_departmentlist(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    departments = department.objects.prefetch_related('designation_set').all()

    return render(request, 'register/view_departmentlist.html', {'departments': departments, 'allocated_menus': allocated_menus})

def update_departmentlist(request, id):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    mem = get_object_or_404(department, id=id)

    if request.method == 'POST':
        form = department_Form(request.POST, instance=mem)
        if form.is_valid():
            form.save()
            return redirect('view_departmentlist')  # Redirect to the list view or any relevant view
    else:
        form = department_Form(instance=mem)

    return render(request, 'register/update_departmentlist.html', {'form': form, 'allocated_menus': allocated_menus})


def delete_departmentlist(request,id):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Fetch the object or return a 404 error if not found
    dtl = get_object_or_404(department,id=id)

    if request.method == "POST":
        dtl.delete()
        return redirect('view_departmentlist')

    # Render the confirmation page for GET requests
    return render(request, 'register/delete_departmentlist.html', {'dtl': dtl,'allocated_menus':allocated_menus})


def view_designation(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    # Sort the designations by department name
    vie = designation.objects.select_related('dept').order_by('dept__dept_Name')
    # Set up pagination
    paginator = Paginator(vie, 10)  # Show 10 designations per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'register/view_designation.html', {'page_obj': page_obj, 'allocated_menus': allocated_menus})


def update_designation(request, id):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    mem = get_object_or_404(designation, id=id)

    if request.method == 'POST':
        form = designation_Form(request.POST, instance=mem)
        if form.is_valid():
            form.save()
            return redirect('view_designation')  # Redirect to the list view or any relevant view
    else:
        form = designation_Form(instance=mem)

    return render(request, 'register/update_designation.html', {'form': form, 'allocated_menus': allocated_menus})

def delete_designation(request,id):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    # Fetch the object or return a 404 error if not found
    dtl = get_object_or_404(designation,id=id)

    if request.method == "POST":
        dtl.delete()
        return redirect('view_designation')

    # Render the confirmation page for GET requests
    return render(request, 'register/delete_designation.html', {'dtl': dtl,'allocated_menus':allocated_menus})


def delete_selected_users(request):
    if request.method == 'POST':
        selected_ids = request.POST.get('selected_ids', '').split(',')
        if not selected_ids or selected_ids == ['']:
            return HttpResponse("No users selected for deletion.", status=400)

        try:
            # Delete users by their IDs
            CustomUser.objects.filter(id__in=selected_ids).delete()
            return redirect('view_users')  # Adjust 'view_users' to the correct URL name
        except Exception as e:
            # Log the error (optional)
            print(f"Error deleting users: {e}")
            return HttpResponse(f"An error occurred: {e}", status=500)
    return HttpResponse("Method not allowed", status=405)

def delete_selected_department(request):
    if request.method == 'POST':
        selected_ids = request.POST.get('selected_ids', '').split(',')
        selected_ids = [id.strip() for id in selected_ids if id.strip()]

        if not selected_ids:
            return HttpResponse("No departments selected for deletion.", status=400)

        try:
            # Ensure the selected IDs are integers
            selected_ids = [int(id) for id in selected_ids]
            # Delete departments by their IDs
            department.objects.filter(id__in=selected_ids).delete()
            return redirect('view_departmentlist')  # Adjust this URL as necessary
        except Exception as e:
            # Log the error (optional)
            print(f"Error deleting departments: {e}")
            return HttpResponse(f"An error occurred: {e}", status=500)

    return HttpResponse("Method not allowed", status=405)


