from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from datetime import date,datetime
from Glenda_App.models import Menu
from logistics_app.forms import DealerForm, DriverForm, VehicleForm, RouteForm, RoutePlanForm, VehicleMaintenanaceForm, DispatchAdviceForm
from logistics_app.models import Driver, Route, Route_Plan, Vehicle, Vehicle_Maintenance
from django.contrib import messages
import csv
from django.db.models import Q

from register_app.models import MenuPermissions
from .models import DispatchAdvice, Privatevehicle, Route, dealer_reg, trackorder
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment



def Add_dealer(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    if request.method == 'POST':
        form = DealerForm(request.POST)
        if form.is_valid():
            form.save()  # Save the driver instance
            return redirect('view_dealer')  # Update with your actual success URL
        else:
            # Print form errors to debug
            print(form.errors)  # You can also use logging or just render errors
    else:
        form = DealerForm()

    return render(request, 'logistics/dealer.html', {'form': form, 'allocated_menus': allocated_menus})

def view_dealer(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)


    dr = dealer_reg.objects.all()
   
    return render(request, 'logistics/viewdealer.html', {'allocated_menus': allocated_menus,'dr': dr})

def Add_driver(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    if request.method == 'POST':
        form = DriverForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()  # Save the driver instance
            return redirect('view_driver')  # Update with your actual success URL
    else:
        form = DriverForm()
    range_5 = range(1, 6) 
    return render(request, 'logistics/add_driver.html', {'form': form, 'allocated_menus': allocated_menus})


def view_driver(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)


    dr = Driver.objects.all()
    today = date.today()
    range_5 = range(1, 6)
    return render(request, 'logistics/view_driver.html', {'range_5': range_5,'allocated_menus': allocated_menus, 'dr': dr})


def update_driver(request, id):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)


    mem = get_object_or_404(Driver, id=id)

    today = date.today()
    license_exp_date = request.POST.get('license_exp_date')

    if request.method == 'POST':
        form = DriverForm(request.POST, instance=mem)
        if form.is_valid():

            license_exp_date = form.cleaned_data['license_exp_date']
            aadhaar_number = form.cleaned_data['aadhaar_number']

            if license_exp_date < today:
                messages.warning(request, "Your Driving License is Expired")
                return render(request, 'logistics/add_driver.html', {'form': form, 'allocated_menus': allocated_menus})

            elif license_exp_date == today:
                messages.warning(request, "Your Driving License expires today")
                return render(request, 'logistics/add_driver.html', {'form': form, 'allocated_menus': allocated_menus})

            if len(str(aadhaar_number)) != 12:
                messages.warning(request, 'Aadhar number must be exactly 12 digits.')
                return render(request, 'logistics/add_driver.html', {'form': form, 'allocated_menus': allocated_menus})

            form.save()
            return redirect('view_driver')  # Redirect to the list view or any relevant view
    else:
        form = DriverForm(instance=mem)

    return render(request, 'logistics/update_driver.html', {'form': form, 'allocated_menus': allocated_menus})


def delete_driver(request, id):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    # Fetch the object or return a 404 error if not found
    dtl = get_object_or_404(Driver, id=id)

    if request.method == "POST":
        dtl.delete()
        return redirect('view_driver')

    # Render the confirmation page for GET requests
    return render(request, 'logistics/delete_driver.html', {'dtl': dtl,'allocated_menus':allocated_menus})


def Add_vehicle(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)


    if request.method == 'POST':
        form = VehicleForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Vehicle created successfully!')
            return redirect('view_vehicle')  # Change to your desired redirect URL
        else:
            # If the form is not valid, errors will be available in `form.errors`
            messages.error(request, 'Please correct the errors below.')
    else:
        form = VehicleForm()
    return render(request, 'logistics/add_vehicle.html', {'form': form, 'allocated_menus': allocated_menus})


def view_vehicle(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    dr = Vehicle.objects.all()
    return render(request, 'logistics/view_vehicle.html', {'allocated_menus': allocated_menus, 'dr': dr})


def update_vehicle(request, id):
    use = request.user  # Get the current user
    vehicle = get_object_or_404(Vehicle, id=id)

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)


    if request.method == 'POST':
        form = VehicleForm(request.POST, request.FILES, instance=vehicle)
        if form.is_valid():
            form.save()  # Save the updated vehicle record
            messages.success(request, 'Vehicle updated successfully!')
            return redirect('view_vehicle')  # Redirect to the vehicle list or another page
    else:
        form = VehicleForm(instance=vehicle)

    return render(request, 'logistics/update_vehicle.html', {'form': form, 'allocated_menus': allocated_menus})


def delete_vehicle(request, id):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Fetch the object or return a 404 error if not found
    dtl = get_object_or_404(Vehicle, id=id)

    if request.method == "POST":
        dtl.delete()
        return redirect('view_vehicle')

    # Render the confirmation page for GET requests
    return render(request, 'logistics/delete_vehicle.html', {'dtl': dtl})




def create_route(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    if request.method == 'POST':
        form = RouteForm(request.POST)
        if form.is_valid():
            # Create and save a new Route instance
            route = Route(
                route_name=form.cleaned_data['route_name'],
                starting_point=form.cleaned_data['starting_point'],
                ending_point=form.cleaned_data['ending_point'],
                total_distance=form.cleaned_data['total_distance']
            )
            route.save()  # Save the instance to the database

            # Redirect to a success page or render a success message
            return redirect('view_route')  # Change this to your actual success URL
        else:
            print(form.errors)  # Print form errors for debugging

    else:
        form = RouteForm()

    return render(request, 'logistics/add_route.html', {
        'form': form,
        'allocated_menus': allocated_menus
    })

from django.http import JsonResponse

def get_route_details(request):
    route_id = request.GET.get('route_id')

    if route_id:
        # Attempt to fetch the route object
        route = get_object_or_404(Route, id=route_id)  # Use get_object_or_404 for cleaner error handling

        data = {
            'total_distance': route.total_distance,  # Field from Route model
            'starting_point': route.starting_point,  # Field from Route model
            'ending_point': route.ending_point,  # Field from Route model
        }
        return JsonResponse(data)

    return JsonResponse({'error': 'No route ID provided'}, status=400)



def view_route(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    dr = Route.objects.all()
    return render(request, 'logistics/view_route.html', {'allocated_menus': allocated_menus, 'dr': dr})


def update_route(request, id):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    mem = get_object_or_404(Route, id=id)

    if request.method == 'POST':
        form = RouteForm(request.POST, instance=mem)
        if form.is_valid():
            form.save()
            return redirect('view_route')  # Redirect to the list view or any relevant view
    else:
        form = RouteForm(instance=mem)

    return render(request, 'logistics/update_route.html', {'form': form, 'allocated_menus': allocated_menus})


def delete_route(request, id):
    # Fetch the object or return a 404 error if not found
    dtl = get_object_or_404(Route, id=id)

    if request.method == "POST":
        dtl.delete()
        return redirect('view_route')

    # Render the confirmation page for GET requests
    return render(request, 'logistics/delete_route.html', {'dtl': dtl})


def Add_routeplan(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    if request.method == 'POST':
        form = RoutePlanForm(request.POST, request.FILES)
        if form.is_valid():
            date = form.cleaned_data['date']
            time = form.cleaned_data['time']
            route = form.cleaned_data['route']
            driver = form.cleaned_data['driver']

            # Check for conflicts based on date and route
            if Route_Plan.objects.filter(date=date, route=route).exists():
                messages.warning(request, "Cannot add to the same route on this date.")
                return render(request, 'logistics/add_routeplan.html', {'form': form, 'allocated_menus': allocated_menus})

            # Check for conflicts based on date and time
            if Route_Plan.objects.filter(date=date, time=time).exists():
                messages.warning(request, "Cannot assign a driver to the same route at this date and time.")
                return render(request, 'logistics/add_routeplan.html', {'form': form, 'allocated_menus': allocated_menus})

            # If no matches are found (neither exact nor conflicting), save the new entry
            else:
                form.save()
                messages.success(request, "Route plan added successfully.")
                return redirect('view_routeplan')

        else:
            # If the form is invalid, re-render with errors
            return render(request, 'logistics/add_routeplan.html', {'form': form, 'allocated_menus': allocated_menus})

    else:
        form = RoutePlanForm()

    # Render the form on initial load or GET request
    return render(request, 'logistics/add_routeplan.html', {'form': form, 'allocated_menus': allocated_menus})




def view_routeplan(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    dr = Route_Plan.objects.all()
    return render(request, 'logistics/view_routeplan.html', {'allocated_menus': allocated_menus, 'dr': dr})


def update_routeplan(request, id):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    mem = get_object_or_404(Route_Plan, id=id)

    if request.method == 'POST':
        form = RoutePlanForm(request.POST, instance=mem)
        if form.is_valid():
            date = form.cleaned_data['date']
            time = form.cleaned_data['time']
            route = form.cleaned_data['route']
            driver = form.cleaned_data['driver']

            # Check for exact match excluding the current instance being updated
            if Route_Plan.objects.exclude(id=id).filter(date=date, time=time, route=route).exists():
                messages.warning(request, "Cannot update to the same route at this date and time.")
                return render(request, 'logistics/update_routeplan.html', {'form': form, 'allocated_menus': allocated_menus})

            elif Route_Plan.objects.exclude(id=id).filter(date=date, route=route).exists():
                messages.warning(request, "Cannot update to the same route at this date and time.")
                return render(request, 'logistics/update_routeplan.html', {'form': form, 'allocated_menus': allocated_menus})

            else:
                form.save()
                messages.success(request, "Route plan updated successfully.")
                return redirect('view_routeplan')

    else:
        form = RoutePlanForm(instance=mem)

    return render(request, 'logistics/update_routeplan.html', {'form': form, 'allocated_menus': allocated_menus})




def delete_routeplan(request, id):
    # Fetch the object or return a 404 error if not found
    dtl = get_object_or_404(Route_Plan, id=id)

    if request.method == "POST":
        dtl.delete()
        return redirect('view_routeplan')

    # Render the confirmation page for GET requests
    return render(request, 'logistics/delete_routeplan.html', {'dtl': dtl})


def driver_search_and_export(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    driver_list = Driver.objects.all()  # Default to all vendors

    filters = Q()  # Initialize filters as an empty Q object

    if request.method == 'GET':
        search_query = request.GET.get('search_query', '').strip()
        export_format = request.GET.get('export', '')

        if search_query:
            # Build filters based on the search query
            if search_query.isdigit():
                filters &= Q(phone_number__icontains=search_query)  # Filter by phone number
            else:
                filters &= Q(user__name__icontains=search_query)
                # Apply filters if any were provided
        if filters:
            driver_list = Driver.objects.filter(filters)

        if 'export' in request.GET:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Driver List'
            ws.merge_cells('A1:D1')
            ws['A1'] = 'Driver List'
            headline_font = Font(bold=True, size=14)
            ws['A1'].font = headline_font
            ws['A1'].alignment = Alignment(horizontal='center')

            headers = ['Name', 'Aadhar Number', 'Phone Number']
            ws.append(headers)

            header_font = Font(bold=True)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_num)
                cell.font = header_font

            for driver in driver_list:
                ws.append([driver.driver_name, driver.aadhaar_number, driver.phone_number])

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="driver_list.xlsx"'

            wb.save(response)

            return response

    context = {
        'dr': driver_list,  # This will be used in the template
        'allocated_menus': allocated_menus,
    }

    return render(request, 'logistics/view_driver.html', context)

def route_search_and_export(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)


    route_list = Route.objects.all()  # Default to all vendors

    filters = Q()  # Initialize filters as an empty Q object

    if request.method == 'GET':
        search_query = request.GET.get('search_query', '').strip()
        export_format = request.GET.get('export', '')

        if search_query:
            # Build filters based on the search query
            if search_query:

                filters &= Q(route_name__icontains=search_query) | Q(starting_point__icontains=search_query) | Q(ending_point__icontains=search_query)
            # Filter by phone number

        # Apply filters if any were provided
        if filters:
            route_list = Route.objects.filter(filters)

        if 'export' in request.GET:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Route List'
            ws.merge_cells('A1:D1')
            ws['A1'] = 'Route List'
            headline_font = Font(bold=True, size=14)
            ws['A1'].font = headline_font
            ws['A1'].alignment = Alignment(horizontal='center')

            headers = ['Name', 'Starting Point', 'Ending point']
            ws.append(headers)

            header_font = Font(bold=True)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_num)
                cell.font = header_font

            for route in route_list:
                ws.append([route.route_name,route.starting_point,route.ending_point])

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="route_list.xlsx"'

            wb.save(response)
            return response

    context = {
        'dr': route_list,  # This will be used in the template
        'allocated_menus': allocated_menus,
    }

    return render(request, 'logistics/view_route.html', context)

def vehicle_search_and_export(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    vehicle_list = Vehicle.objects.all()  # Default to all vendors

    filters = Q()  # Initialize filters as an empty Q object

    if request.method == 'GET':
        search_query = request.GET.get('search_query', '').strip()
        export_format = request.GET.get('export', '')


        if search_query:

            filters &= Q(vehicle_name__icontains=search_query) | Q(vehicle_nbr__icontains=search_query)
            # Filter by phone number

        # Apply filters if any were provided
        if filters:
            vehicle_list = Vehicle.objects.filter(filters)

        if 'export' in request.GET:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Vehicle List'
            ws.merge_cells('A1:D1')
            ws['A1'] = 'Vehicle List'
            headline_font = Font(bold=True, size=14)
            ws['A1'].font = headline_font
            ws['A1'].alignment = Alignment(horizontal='center')

            headers = ['Name', 'Number', 'Fasttag','Taxes_date','Pollution_date','Insurance_date']
            ws.append(headers)

            header_font = Font(bold=True)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_num)
                cell.font = header_font

            for vehicle in vehicle_list:
                ws.append([vehicle.vehicle_name,vehicle.vehicle_nbr,vehicle.vehicle_fasttag,vehicle.vehicle_taxes_date,vehicle.vehicle_pollution_date,vehicle.vehicle_insurance_date])

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="vehicle_list.xlsx"'

            wb.save(response)
            return response

    context = {
        'dr': vehicle_list,  # This will be used in the template
        'allocated_menus': allocated_menus,
    }

    return render(request, 'logistics/view_vehicle.html', context)

def route_plan_search_and_export(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    route_plan_list = Route_Plan.objects.all()  # Default to all route plans

    filters = Q()  # Initialize filters as an empty Q object

    if request.method == 'GET':
        search_query = request.GET.get('search_query', '').strip()
        export_format = request.GET.get('export', '')

        if search_query:
            # Modify to search by related fields in the Vehicle and Driver models
            filters &= Q(vehicle__vehicle_name__icontains=search_query) | Q(driver__driver_name__icontains=search_query)

        # Apply filters if any were provided
        if filters:
            route_plan_list = Route_Plan.objects.filter(filters)

        if 'export' in request.GET:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Route Plan List'
            ws.merge_cells('A1:D1')
            ws['A1'] = 'Route Plan List'
            headline_font = Font(bold=True, size=14)
            ws['A1'].font = headline_font
            ws['A1'].alignment = Alignment(horizontal='center')

            headers = ['Vehicle Name', 'Driver Name', 'Date', 'Time']
            ws.append(headers)

            header_font = Font(bold=True)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_num)
                cell.font = header_font

            for route in route_plan_list:
                ws.append([route.vehicle.vehicle_name, route.driver.driver_name, route.date, route.time])

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="route_plan_list.xlsx"'

            wb.save(response)
            return response

    context = {
        'dr': route_plan_list,  # This will be used in the template
        'allocated_menus': allocated_menus,
    }

    return render(request, 'logistics/view_routeplan.html', context)


def VehicleMaintenanace(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    if request.method == 'POST':
        form = VehicleMaintenanaceForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('Maintenanace')
    else:
        form = VehicleMaintenanaceForm()
    return render(request, 'logistics/vehicle_maintanance.html', {'form': form, 'allocated_menus': allocated_menus})


def view_vmaintenance(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    dr = Vehicle_Maintenance.objects.all()
    return render(request, 'logistics/view_vehiclemaint.html', {'allocated_menus': allocated_menus, 'dr': dr})


def update_vmaintenance(request, id):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    mem = get_object_or_404(Vehicle_Maintenance, id=id)

    if request.method == 'POST':
        form = VehicleMaintenanaceForm(request.POST, instance=mem)
        if form.is_valid():
            form.save()
            return redirect('view_vmaintenance')  # Redirect to the list view or any relevant view
    else:
        form = VehicleMaintenanaceForm(instance=mem)

    return render(request, 'logistics/update_vmaintenance.html', {'form': form, 'allocated_menus': allocated_menus})


def delete_vmaintenance(request, id):
    # Fetch the object or return a 404 error if not found
    dtl = get_object_or_404(Vehicle_Maintenance, id=id)

    if request.method == "POST":
        dtl.delete()
        return redirect('view_vmaintenance')

    # Render the confirmation page for GET requests
    return render(request, 'logistics/delete_vmaintenanace.html', {'dtl': dtl})


def vmaintenance_search(request):
    filters = {}
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # if request.method == 'GET':
    search_query = request.GET.get('search_query', '').strip()

    if search_query:
        maintenance_list = Vehicle_Maintenance.objects.all()
    else:
        filters = Q()
        if search_query:
            filters &= Q(vehicle__icontains=search_query)
        # Apply filters if any were provided
        maintenance_list = Vehicle_Maintenance.objects.filter(filters)

    context = {
        'view': maintenance_list,  # This will be used in the template
        'allocated_menus': allocated_menus,
        'search_query': search_query,
    }

    return render(request, 'logistics/view_vehiclemaint.html', context)


def view_vmaintenancesearch(request, id):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    view = Vehicle_Maintenance.objects.all()
    print(view)

    return render(request, 'logistics/vmaintenance_search.html', {'view': view, 'allocated_menus': allocated_menus})


def maintenance_csv(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    view = Vehicle_Maintenance.objects.all()

    context = {'view': view, 'menu': allocated_menus}
    response = HttpResponse(content_type='application/csv')
    response['Content-Disposition'] = f'attachment; filename="maintenance.csv"'

    writer = csv.writer(response)
    writer.writerow(['Vehicle Name'])

    vmaintenance = Vehicle_Maintenance.objects.all()

    for vehicle in vmaintenance:
        writer.writerow([vehicle.vehicle])

    return response


def vmaintenance_exportsearch(request):
    # Prefetch allocated_menus for rendering in the template
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)
    export_format = request.GET.get('export', '')
    search_query = request.GET.get('search_query', '').strip()  # Search by name or phone number

    # Initialize the filters
    filters = Q()

    # Initialize the export list based on the search query
    if search_query:
        # Check if the search query is a digit (indicating a phone number)
        if search_query:
            filters &= Q(vehicle__icontains=search_query)

    # Fetch the filtered list based on the constructed filters
    searchexport_list = Vehicle_Maintenance.objects.filter(filters)

    # If 'export' parameter is found, handle CSV export
    if export_format == 'excel':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="Customer_searchexport_List.csv"'

        writer = csv.writer(response)
        writer.writerow(['vehicle', 'TypeofMaintenance', 'amount', 'Remarks'])

        for vehicle in searchexport_list:
            writer.writerow([
                vehicle.vehicle,
                vehicle.TypeofMaintenance,
                vehicle.amount,
                vehicle.Remarks

            ])

        return response  # Return CSV response for download

    # Prepare the context to render in the template
    context = {
        'view': searchexport_list,  # Filtered customer list for display
        'allocated_menus': allocated_menus,  # Menus for navigation
        'search_query': search_query,  # Keep search query in the context for display in form
    }

    # Render the combined view template
    return render(request, 'logistics/view_vehiclemaint.html', context)



def vehicle_maintenance_search_and_export(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    main_list = Vehicle_Maintenance.objects.all()  # Default to all vendors
    filters = Q()  # Initialize filters as an empty Q object

    if request.method == 'GET':
        search_query = request.GET.get('search_query', '').strip()
        export_format = request.GET.get('export', '')

        if search_query:
            # Build filters based on the search query
            filters |= Q(vehicle__vehicle_name__icontains=search_query)  # Filter by vehicle name
            filters |= Q(TypeofMaintenance__icontains=search_query)  # Filter by type of maintenance
            filters |= Q(vehicle__vehicle_nbr__icontains=search_query)  # Filter by vehicle name

        # Apply filters if any were provided
        if filters:
            main_list = Vehicle_Maintenance.objects.filter(filters)

        if export_format:  # Check if export is requested
            wb = Workbook()
            ws = wb.active
            ws.title = 'Vehicle Maintenance List'
            ws.merge_cells('A1:D1')
            ws['A1'] = 'Vehicle Maintenance List'
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')

            headers = ['Vehicle Name', 'Type of Maintenance', 'Amount']
            ws.append(headers)

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_num)
                cell.font = Font(bold=True)

            for vehicle in main_list:
                ws.append([vehicle.vehicle.vehicle_name, vehicle.TypeofMaintenance, vehicle.amount])

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="maintenance_list.xlsx"'

            wb.save(response)
            return response

    context = {
        'dr': main_list,  # This will be used in the template
        'allocated_menus': allocated_menus,
    }

    # Render the same template with updated context
    return render(request, 'logistics/view_vehiclemaint.html', context)

def Dispatchadvice(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    if request.method == 'POST':
        form = DispatchAdviceForm(request.POST)
        if form.is_valid():
            form.save()  # Save the driver instance
            return redirect('view_dispatchadvice')  # Update with your actual success URL
        else:
            # Print form errors to debug
            print(form.errors)  # You can also use logging or just render errors
    else:
        form = DispatchAdviceForm()

    return render(request, 'logistics/add_dispatch.html', {'form': form, 'allocated_menus': allocated_menus})

def view_dispatchadvice(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)


    dr = DispatchAdvice.objects.all()
   
    return render(request, 'logistics/view_dispatch.html', {'allocated_menus': allocated_menus,'dr': dr})

# 
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from django.shortcuts import render

def dispatch_search_and_export(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    dispatch_list = DispatchAdvice.objects.all()  # Default to all dispatches
    filters = Q()  # Initialize filters as an empty Q object

    if request.method == 'GET':
        search_query = request.GET.get('search_query', '').strip()
        export_format = request.GET.get('export', '')

        if search_query:
            # Test individual filters to ensure they are valid
            if 'water_Finished_Goods' in DispatchAdvice._meta.get_fields():
                filters |= Q(water_Finished_Goods__name__icontains=search_query)
            if 'Route_Plan' in DispatchAdvice._meta.get_fields():
                filters |= Q(Route_Plan__route_plan__icontains=search_query)
            if 'Dealer' in DispatchAdvice._meta.get_fields():
                filters |= Q(Dealer__dealer__icontains=search_query)
            if 'Dealer' in DispatchAdvice._meta.get_fields():
                filters |= Q(Dealer__dealer__icontains=search_query)



        # Apply filters if any were provided
        if filters:
            dispatch_list = DispatchAdvice.objects.filter(filters)

        if export_format:  # Check if export is requested
            wb = Workbook()
            ws = wb.active
            ws.title = 'Dispatch_list'
            ws.merge_cells('A1:D1')
            ws['A1'] = 'Dispatch List'
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')

            headers = ['Product Name', 'Route Plan', 'Dealer']
            ws.append(headers)

            # Add header style
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_num)
                cell.font = Font(bold=True)

            # Populate the worksheet with data from the dispatch list
            for dispatch in dispatch_list:
                # Make sure you reference the correct fields for product name, route plan, and dealer
                ws.append([dispatch.water_Finished_Goods.name, dispatch.Route_Plan.route_plan, dispatch.Dealer.dealer])

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="Dispatch List.xlsx"'

            wb.save(response)
            return response

    context = {
        'dr': dispatch_list,
        'allocated_menus': allocated_menus,
    }

    # Render the same template with updated context
    return render(request, 'logistics/view_dispatch.html', context)

def Add_privatevehicle(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    if request.method == 'POST':
        form = PrivatevehicleForm(request.POST)
        if form.is_valid():
            form.save()  # Save the driver instance
            return redirect('view_privatevehicle')  # Update with your actual success URL
        else:
            # Print form errors to debug
            print(form.errors)  # You can also use logging or just render errors
    else:
        form = PrivatevehicleForm()

    return render(request, 'logistics/add_pvtv.html', {'form': form, 'allocated_menus': allocated_menus})

def view_privatevehicle(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)


    dr = Privatevehicle.objects.all()
   
    return render(request, 'logistics/view_pvtv.html', {'allocated_menus': allocated_menus,'dr': dr})


from django.shortcuts import render, redirect
from .forms import PrivatevehicleForm, trackorderForm  # Your form class
from .models import dealer_reg

def trackorder_create(request):
    if request.method == 'POST':
        form = trackorderForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('view_trackorder')  # Redirect to the list page after saving
    else:
        form = trackorderForm()

    dealers = dealer_reg.objects.all()  # Get all dealers from the database
    return render(request, 'logistics/createtrackorder.html', {'form': form, 'dealers': dealers})


def view_trackorder(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)


    dr = trackorder.objects.all()
   
    return render(request, 'logistics/viewtrack.html', {'allocated_menus': allocated_menus,'dr': dr})

