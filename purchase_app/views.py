import datetime
import random
from itertools import chain

from django.shortcuts import render, redirect, get_object_or_404
from Glenda_App.models import Menu
from purchase_app.forms import CategoryForm, RawMaterialForm, RFQRawMaterialsForm
from inventory_app.forms import Raw_materials_StockForm
from django.contrib import messages
from django.db.models import Q
from purchase_app.models import RawMaterials, RawMaterialCategory, RFQ_raw_materials, PurchaseOrder
from inventory_app.models import Raw_material_request, RawMaterialsStock, Purchase_request_raw_material, \
    Purchase_request_semi_finished
from rd_app.models import RD_stock
from register_app.models import MenuPermissions
from django.utils.dateparse import parse_date
from django.db.models.functions import Cast
from django.db.models import DateField
from django.template.loader import get_template
from django.http import HttpResponse, JsonResponse
from xhtml2pdf import pisa
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
from django.core.paginator import Paginator

from vendor_app.models import vendor_quotation


# Create your views here.


def view_rawmaterials(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    view_ca=RawMaterialCategory.objects.all()
    view=RawMaterials.objects.all()
    categories = RawMaterialCategory.objects.all()

    # Pagination logic
    paginator = Paginator(view, 5)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request,'purchase/view_rawmaterials.html',{'view':page_obj,'view_ca':view_ca,'allocated_menus': allocated_menus, 'categories': categories})


def add_category(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    form = CategoryForm()
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        print("Request")
        if form.is_valid():
            print("Success")
            form.save()
            messages.success(request, 'Successfull')

            return redirect('view_rawmaterials')
    return render(request, 'purchase/Add_category.html', {'form': form, 'allocated_menus': allocated_menus})


def create_raw_material(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    if request.method == 'POST':
        form = RawMaterialForm(request.POST, request.FILES)  # Use request.FILES for file upload
        if form.is_valid():
            form.save()
            return redirect('view_rawmaterials')  # Redirect to a list view or another page after saving
    else:
        form = RawMaterialForm()

    return render(request, 'purchase/add_rawmaterials.html', {'form': form, 'allocated_menus': allocated_menus})

def update_rawmaterials(request, id):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    mem = get_object_or_404(RawMaterials, id=id)

    if request.method == 'POST':
        form = RawMaterialForm(request.POST,request.FILES,instance=mem)
        if form.is_valid():
            form.save()
            return redirect('view_rawmaterials')  # Redirect to the list view or any relevant view
    else:
        form = RawMaterialForm(instance=mem)

    return render(request, 'purchase/update_rawmaterials.html', {'form': form, 'allocated_menus': allocated_menus})


def delete_rawmaterils(request, id):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Fetch the object or return a 404 error if not found
    dtl = get_object_or_404(RawMaterials, id=id)

    if request.method == "POST":
        dtl.delete()
        return redirect('view_rawmaterials')

    # Render the confirmation page for GET requests
    return render(request, 'purchase/delete_rawmaterials.html', {'dtl': dtl,'allocated_menus': allocated_menus})


def message_request(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    message = Raw_material_request.objects.filter(status='Pending').order_by('-date')

    # Pagination logic
    paginator = Paginator(message, 5)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request,'purchase/message_request_list.html', {'message':page_obj, 'allocated_menus': allocated_menus})


def message_response(request, id):
    # Fetch menus and counts
    use = request.user
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')
    allocated_menus = {perm.menu: [] for perm in user_permissions}
    for perm in user_permissions:
        allocated_menus[perm.menu].append(perm.sub_menu)

    request_data = get_object_or_404(Raw_material_request, pk=id)

    if request.method == 'POST':
        action_type = request.POST.get('action_type')
        response_text = request.POST.get('response')

        if action_type == 'accept':
            request_data.status = 'Accepted'
            request_data.response = response_text
            request_data.save()
            return redirect('message_requests')

        elif action_type == 'decline':
            if response_text:
                request_data.status = 'Declined'
                request_data.response = response_text
                request_data.save()
                return redirect('message_requests')
            else:
                error_message = "Please provide a reason for declining"
                return render(request, 'purchase/message_request_view.html', {
                    'data': request_data,
                    'allocated_menus': allocated_menus,
                    'error_message': error_message
                })

    return render(request, 'purchase/message_request_view.html', {
        'data': request_data,
        'allocated_menus': allocated_menus
    })


def raw_material_purchase_search(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    categories = RawMaterialCategory.objects.all()  # Get distinct categories
    raw_materials = RawMaterials.objects.prefetch_related('stocks').all()

    context = {
        'view': raw_materials,
        'categories': categories,
        'allocated_menus': allocated_menus
    }

    if request.method == 'POST':
        category = request.POST.get('category', None)  # Get category from form
        name = request.POST.get('name', None)          # Get name from form

        filters = Q()

        # Apply category filter if selected
        if category and category.isdigit():
            filters &= Q(category_id=int(category))

        # Apply name filter if provided
        if name:
            filters &= Q(name__icontains=name)

        # Filter stock based on combined filters
        if filters:
            search_list = RawMaterials.objects.filter(filters)

            # Debugging: Logging or print statement
            print(f"Filters applied: {filters}")
            print(f"Filtered records count: {search_list.count()}")

    context['view'] = search_list # Filtered stock data

    return render(request, 'purchase/view_rawmaterials.html', context)


def add_stocks(request, id):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    raw_material = get_object_or_404(RawMaterials, id=id)

    if request.method == 'POST':
        form = Raw_materials_StockForm(request.POST)
        if form.is_valid():
            stock_entry = form.save(commit=False)
            stock_entry.raw_materials = raw_material  # Set the raw material
            stock_entry.status = 'Pending'
            stock_entry.save()  # Save the new stock entry

            # Update the total stock for the raw material by adding new stock
            # new_stock = stock_entry.stock
            # raw_material.total_stock = F('total_stock') + new_stock
            # raw_material.save(update_fields=["total_stock"])  # Save the updated total stock

            return redirect('view_rawmaterials')  # Redirect to the list view
    else:
        form = Raw_materials_StockForm()

    return render(request, 'purchase/add_raw_materials_stock.html', {
        'form': form,
        'allocated_menus': allocated_menus,
        'raw_material': raw_material,
    })


def rawmaterials_stock_history(request,id):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Retrieve all RawMaterialsStock instances with the given raw_materials_id
    raw_materials_stock_qs = RawMaterialsStock.objects.filter(raw_materials_id=id)


    # Filter RD_stock entries associated with the retrieved RawMaterialsStock instances
    raw_materials = RD_stock.objects.filter(raw_materials__in=raw_materials_stock_qs).select_related(
        'raw_materials', 'raw_materials__raw_materials', 'raw_materials__raw_materials__category'
    ).order_by('-date', '-time')

    # Date-based filtering
    filter_type = request.GET.get('filter')
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parse_date_start = parse_date(query_date) if query_date else None
    parse_date_end = parse_date(query_date_end) if query_date_end else None
    is_filtered = False

    # Apply date range filtering if both start and end dates are provided
    if parse_date_start and parse_date_end:
        raw_materials = raw_materials.annotate(date_only=Cast('date', DateField())).filter(date_only__range=[parse_date_start, parse_date_end])
        is_filtered = True
    elif parse_date_start:
        raw_materials = raw_materials.annotate(date_only=Cast('date', DateField())).filter(date_only=parse_date_start)
        is_filtered = True

    # Pagination logic
    paginator = Paginator(raw_materials, 7)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Pass the filtered results and query_date to the template
    context = {
        'data': page_obj,
        'allocated_menus': allocated_menus,
        'id': id,
        'filter_type': filter_type,
        'query_date': query_date,
        'query_date_end': query_date_end,
        'is_filtered': is_filtered
    }

    return render(request, 'purchase/view_raw_materials_history.html', context)


def message_request_list_history(request):
    # Fetch menus and counts
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    message = Raw_material_request.objects.all().order_by('-date')

    # Retrieve and validate the query date parameter
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date_start = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date_end else None
    is_filtered = False

    # Apply date filtering if parsed_date is valid
    if parsed_date_start and parsed_date_end:
        message = message.annotate(date_only=Cast('date', DateField())).filter(
            date_only__range=[parsed_date_start, parsed_date_end])
        is_filtered = True

    elif parsed_date_start:
        message = message.annotate(date_only=Cast('date', DateField())).filter(
            date_only=parsed_date_start
        )
        is_filtered = True

    # Pagination logic
    paginator = Paginator(message, 7)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'allocated_menus': allocated_menus,
        'message': page_obj,
        'is_filtered': is_filtered,
        'query_date': query_date,
        'query_date_end': query_date_end,
    }

    return render(request,'purchase/message_request_list_history.html',context)


def raw_message_request_list_history_pdf(request):
    # Retrieve and validate the query date parameter
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date else None

    # Retrieve filtered or unfiltered data based on query date
    data = Raw_material_request.objects.all()
    if parsed_date and parsed_date_end:
        data = data.annotate(date_only=Cast('date', DateField())).filter(date_only__range=[parsed_date,parsed_date_end])
    elif parsed_date:
        data = data.annotate(date_only=Cast('date', DateField())).filter(date_only=parsed_date)

    # Retrieve the first entry for dynamic filename generation
    view = data.first()
    filename = f"{view.category.category_name}_request-review_history.pdf" if view else "review-history.pdf"

    # Prepare context for rendering PDF
    context = {'view': data, 'name': view}

    # Load and render the template to HTML
    template = get_template('purchase/pdf_raw_message_request_list_history.html')
    html = template.render(context)

    # Create the HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Generate the PDF
    pdf_status = pisa.CreatePDF(html, dest=response)

    return response


def raw_message_request_list_history_excel(request):
    # Retrieve and validate the query date parameters
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date_end else None

    # Retrieve filtered or unfiltered items based on the query date
    items = Raw_material_request.objects.all()
    if parsed_date and parsed_date_end:
        items = items.annotate(date_only=Cast('date', DateField())).filter(date_only__range=(parsed_date, parsed_date_end))
    elif parsed_date:
        items = items.annotate(date_only=Cast('date', DateField())).filter(date_only=parsed_date)

    # Generate filename based on query date parameters
    filename = f"Raw_Materials_Request_{query_date}to{query_date_end}.xlsx" if query_date and query_date_end else "Raw_Materials_Request.xlsx"

    # Create an Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Materials Request History"

    # Write the title row at the top
    title = f"Raw Materials Request History {query_date} to {query_date_end}" if query_date and query_date_end else "Raw Materials Request History"
    ws.append([title])
    ws.append([])  # Blank row for line space
    ws.append([])  # Second blank row for line space

    # Write the header row
    headers = ['Category', 'Status', 'Date and Time']
    ws.append(headers)

    # Set header row font to bold
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)  # Header is in row 4 after title rows
        cell.font = header_font

    # Iterate through the filtered data and write each row to the Excel file
    for item in items:
        # Convert date to naive format and format as string for Excel compatibility
        date_value = item.date.replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S') if item.date else 'N/A'
        ws.append([
            item.category.category_name if item.category else 'N/A',
            item.status,
            date_value,
        ])

    # Prepare the response as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save the workbook to the response
    wb.save(response)

    return response


def raw_materials_purchase_history_pdf(request,id):
    # Retrieve and validate the query date parameter
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date else None

    # Retrieve all RawMaterialsStock instances with the given raw_materials_id
    raw_materials_stock_qs = RawMaterialsStock.objects.filter(raw_materials_id=id)

    # Filter RD_stock entries associated with the retrieved RawMaterialsStock instances
    data = RD_stock.objects.filter(raw_materials__in=raw_materials_stock_qs).select_related(
        'raw_materials', 'raw_materials__raw_materials', 'raw_materials__raw_materials__category'
    ).order_by('-date', '-time')


    if parsed_date and parsed_date_end:
        data = data.annotate(date_only=Cast('date', DateField())).filter(date_only__range=[parsed_date,parsed_date_end])
    elif parsed_date:
        data = data.annotate(date_only=Cast('date', DateField())).filter(date_only=parsed_date)

    # Retrieve the first entry for dynamic filename generation
    view = data.first()
    filename = f"{view.raw_materials.raw_materials.category.category_name}-purchase history.pdf" if view else "purchase-history.pdf"

    # Prepare context for rendering PDF
    context = {'view': data, 'name': view}

    # Load and render the template to HTML
    template = get_template('purchase/pdf_view_raw_materials_history.html')
    html = template.render(context)

    # Create the HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Generate the PDF
    pdf_status = pisa.CreatePDF(html, dest=response)

    return response


def raw_materials_purchase_history_excel(request, id):
    # Retrieve and validate the query date parameter
    query_date = request.GET.get('query')
    query_date_end = request.GET.get('query_end')
    parsed_date = parse_date(query_date) if query_date else None
    parsed_date_end = parse_date(query_date_end) if query_date_end else None

    # Retrieve all RawMaterialsStock instances with the given raw_materials_id
    raw_materials_stock_qs = RawMaterialsStock.objects.filter(raw_materials_id=id)

    # Filter RD_stock entries associated with the retrieved RawMaterialsStock instances
    raw_materials = RD_stock.objects.filter(raw_materials__in=raw_materials_stock_qs).select_related(
        'raw_materials', 'raw_materials__raw_materials', 'raw_materials__raw_materials__category'
    ).order_by('-date', '-time')

    # Apply date filter only if both parsed dates are valid
    if parsed_date and parsed_date_end:
        raw_materials = raw_materials.filter(date__range=(parsed_date, parsed_date_end))
    elif parsed_date:
        raw_materials = raw_materials.filter(date=parsed_date)

    # Get the raw material name for the filename
    data = raw_materials.first()  # Use the first entry to get material info
    filename = f"{data.raw_materials.raw_materials.category.category_name}_purchase_report.xlsx" if data else "raw_material_purchase_report.xlsx"

    # Create an Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Material Purchase History"

    # Write the headline row that spans all columns
    ws.merge_cells('A1:H1')
    headline = f"{data.raw_materials.raw_materials.name} Raw Material Purchase History" if data else "Raw Material Purchase History"
    ws['A1'] = headline

    # Set font properties and center the headline
    headline_font = Font(bold=True, size=14)
    ws['A1'].font = headline_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # Write the header row
    headers = ['Category', 'Name', 'Size', 'Requested Stock', 'Approved Stock', 'Unapproved Stock', 'Requested Date', 'Approved Date']
    ws.append([])  # Blank row for spacing
    ws.append(headers)

    # Make the header row bold
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)  # Header is in row 3
        cell.font = header_font

    # Iterate through the data and write each row to the Excel file
    for item in raw_materials:
        ws.append([
            item.raw_materials.raw_materials.category.category_name if item.raw_materials.raw_materials.category else 'N/A',  # Category name
            item.raw_materials.raw_materials.name,  # Name of the raw material
            item.raw_materials.raw_materials.size,  # Size of the raw material
            item.raw_materials.stock,
            item.approved_stock,  # Stock amount
            item.unapproved_stock,
            item.raw_materials.raw_materials.date.strftime('%Y-%m-%d') if item.raw_materials.raw_materials.date else 'N/A',
            item.date.strftime('%Y-%m-%d') if item.date else 'N/A'  # Formatting date
        ])

    # Prepare the response as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save the workbook to the response
    wb.save(response)

    return response


def demo_rfq(request):
    # Fetch menus and counts
    use = request.user

    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    form = RFQRawMaterialsForm()
    rfq_requests=RFQ_raw_materials.objects.all()

    return render(request, 'purchase/demo_rfq.html', {'allocated_menus': allocated_menus,'form':form,'rfq':rfq_requests})


def demo_quotations(request):
    use = request.user

    # Fetch user permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Fetch all vendor quotations
    quotations = vendor_quotation.objects.select_related('category').all()

    return render(request, 'purchase/demo_quotations.html', {
        'allocated_menus': allocated_menus,
        'quotations': quotations,  # Pass quotations to the template
    })


def update_status(request, quotation_id, status):
    """
    Updates the status of a vendor quotation.

    Args:
        request: The HTTP request object.
        quotation_id: The ID of the quotation to update.
        status: The new status ('Accept' or 'Reject').

    Returns:
        Redirects to the quotations list page.
    """
    # Fetch the quotation by ID
    quotation = get_object_or_404(vendor_quotation, id=quotation_id)

    # Update the status
    quotation.status = status
    quotation.save()

    # Redirect to the quotations list page
    return redirect('demo_quotations')  # Replace with the name of your quotations list view


def create_po(request, quotation_id):
    if request.method == "POST":
        special_instructions = request.POST.get('special_instructions', '')

        try:
            # Fetch the related quotation data
            quotation = vendor_quotation.objects.get(id=quotation_id)

            # Generate a unique PO number
            today = datetime.date.today().strftime('%Y%m%d')
            po_number = f"PO-{today}-{quotation_id}"

            # Create and save the PurchaseOrder instance
            purchase_order = PurchaseOrder.objects.create(
                quotation_data=quotation,
                po_number=po_number,
                verified_by="Pending",
                special_instructions=special_instructions,
                po_status="Pending"
            )

            # Update the quotation status
            quotation.status = "PO Created"
            quotation.save()

            messages.success(request, f"Purchase Order {purchase_order.po_number} created successfully!")
            return redirect('demo_quotations')

        except vendor_quotation.DoesNotExist:
            messages.error(request, "Quotation not found!")
            return redirect('demo_quotations')
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
            return redirect('demo_quotations')
    else:
        messages.error(request, "Invalid Request!")
        return redirect('demo_quotations')


def demo_purchase_order_list(request):
    use = request.user

    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    return render(request, 'purchase/demo_purchase_order_list.html', {'allocated_menus': allocated_menus})


def demo_purchase_order(request):
    use = request.user

    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    return render(request, 'purchase/demo_purchase_order.html', {'allocated_menus': allocated_menus})


def demo_dispatch_notification(request):
    use = request.user

    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    return render(request, 'purchase/demo_dispatch_notification.html', {'allocated_menus': allocated_menus})


def demo_vendor_list(request):
    use = request.user

    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    return render(request, 'purchase/demo_vendor_list.html', {'allocated_menus': allocated_menus})


def demo_add_vendor(request):
    use = request.user

    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    return render(request, 'purchase/demo_add_vendor.html', {'allocated_menus': allocated_menus})


def demo_all_history(request):
    use = request.user

    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    return render(request, 'purchase/demo_all_history.html', {'allocated_menus': allocated_menus})


def demo_quotation_verification_history(request):
    use = request.user

    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    return render(request, 'purchase/demo_quotation_verification_history.html', {'allocated_menus': allocated_menus})


def demo_add_vendor_history(request):
    use = request.user

    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    return render(request, 'purchase/demo_add_vendor_history.html', {'allocated_menus': allocated_menus})


def demo_rfq_history(request):
    use = request.user

    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    return render(request, 'purchase/demo_rfq_history.html', {'allocated_menus': allocated_menus})


def demo_purchase_order_history(request):
    use = request.user

    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    return render(request, 'purchase/demo_purchase_order_history.html', {'allocated_menus': allocated_menus})


def demo_dispatch_notification_history(request):
    use = request.user

    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    return render(request, 'purchase/demo_dispatch_notification_history.html', {'allocated_menus': allocated_menus})


def demo_new_rfq(request):
    use = request.user

    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    return render(request, 'purchase/demo_new_rfq.html', {'allocated_menus': allocated_menus})


def demo_inventory_request(request):
    use = request.user

    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    raw_requests = Purchase_request_raw_material.objects.all()
    semi_requests = Purchase_request_semi_finished.objects.all()
    # Combine raw_requests and semi_requests
    combined_requests = chain(raw_requests, semi_requests)

    return render(request, 'purchase/demo_inventory_request.html', {'allocated_menus': allocated_menus,'requests':combined_requests})


def demo_vendor_year_report(request):
    use = request.user

    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    return render(request, 'purchase/demo_vendor_year_report.html', {'allocated_menus': allocated_menus})


def demo_req_from_inventory_report(request):
    use = request.user

    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    raw_requests = Purchase_request_raw_material.objects.all()
    semi_requests = Purchase_request_semi_finished.objects.all()
    # Combine raw_requests and semi_requests
    combined_requests = chain(raw_requests, semi_requests)

    return render(request, 'purchase/demo_req_from_inventory_report.html', {'allocated_menus': allocated_menus,'requests':combined_requests})


def create_rfq_raw_materials(request):
    if request.method == 'POST':
        form = RFQRawMaterialsForm(request.POST, request.FILES)

        # Generate a random number for RFQ
        random_number = random.randint(0, 9999999)
        result = f"RFQ{random_number}"

        if form.is_valid():
            rfq_instance = form.save(commit=False)
            rfq_instance.rfq_number = result
            rfq_instance.status = "Sent"
            rfq_instance.save()
            # Save the selected vendors to the many-to-many relationship
            rfq_instance.vendor_list.set(form.cleaned_data['vendor_list'])
            return redirect('demo_rfq')  # Redirect to a list or detail page
        else:
            print(form.errors)
    else:
        form = RFQRawMaterialsForm()

    return render(request, 'purchase/demo_rfq.html', {'form': form})